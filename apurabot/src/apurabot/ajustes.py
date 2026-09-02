"""Os ajustes da apuração — o que a apuração decide e o Livro não sabe.

O Livro Fiscal traz os documentos da competência. As linhas 002, 003, 006 e 007
do Registro de Apuração não vêm de documento nenhum: são decisões tomadas e
aprovadas pelo time fiscal, e por isso só podem chegar declaradas.

Elas chegam por duas portas, que somam na mesma linha do Registro:

**A linha do Livro.** Quando o ajuste tem dono — "esta nota foi lançada errada",
"esta entrada não devia ter crédito" —, ele é informado nas colunas `ajuste_*`
da BASE TRATADA, na própria linha. É a porta preferida, porque o estabelecimento
e a atividade saem da linha: ninguém precisa digitá-los, e ninguém os erra.

**A aba AJUSTES.** Quando o ajuste não tem dono — uma parcela do Registro que
não pertence a nota alguma —, não há linha onde pendurá-lo. Vai para a aba, com
o estabelecimento escrito à mão.

O valor é **sempre positivo**. Quem dá o sentido é a linha escolhida, como no
próprio Registro, onde nenhuma linha aceita número negativo:

    002 Outros Débitos        aumenta o que se deve
    003 Estornos de Créditos  aumenta o que se deve
    006 Outros Créditos       diminui o que se deve
    007 Estornos de Débitos   diminui o que se deve

Reduzir um estorno que a regra calculou não é lançar negativo na 003: é lançar
positivo na 006, que é como o livro escreve isso.

`ANOTAR` é a exceção que não movimenta nada. Serve para marcar uma linha —
"este ICMS é indevido, será tratado por anuência" — sem alterar a apuração. A
linha sai listada em "Marcado, não lançado", com o total, para que o valor em
discussão fique visível em vez de sumir.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from .formato import numero as _numero

#: Linha do Registro → campo de `AjustesDaApuracao` que ela alimenta.
LINHA_DO_REGISTRO = {
    2: "outros_debitos",
    3: "estorno_de_credito",
    6: "outros_creditos",
    7: "estorno_de_debito",
}

#: Marca a linha sem lançar nada no Registro.
ANOTAR = "ANOTAR"

#: Nome da aba onde vão as parcelas sem documento e a conferência.
ABA = "AJUSTES"

TITULO_PARCELAS = "PARCELAS SEM DOCUMENTO"
TITULO_CONFERENCIA = "CONFERÊNCIA"


@dataclass
class Ajuste:
    """Um lançamento declarado, já validado."""

    estabelecimento: str
    #: 2, 3, 6 ou 7 — `None` quando é anotação.
    linha: int | None
    valor: float
    motivo: str
    responsavel: str = ""
    aprovador: str = ""
    #: Preenchida pela apuração quando o ajuste vem de uma linha do Livro.
    atividade: str = ""
    #: Onde ele foi informado, para a memória de cálculo.
    onde: str = ""

    @property
    def anotacao(self) -> bool:
        return self.linha is None

    @property
    def campo(self) -> str:
        return LINHA_DO_REGISTRO[self.linha] if self.linha else ""

    def __str__(self) -> str:
        from .formato import reais

        alvo = "marcado, não lançado" if self.anotacao else f"linha {self.linha:03d}"
        return f"{self.estabelecimento}: {reais(self.valor)} — {alvo} · {self.motivo}"


@dataclass
class Conferencia:
    """A declaração de que um estabelecimento foi conferido.

    Célula vazia diz duas coisas ao mesmo tempo — "não tem ajuste" e "ninguém
    olhou ainda" — e a ferramenta não pode escolher uma. Por isso alguém assina:
    estabelecimento conferido para de esperar ajuste, mesmo sem nenhum.
    """

    estabelecimento: str
    por: str
    em: str = ""
    observacao: str = ""


class AjusteInvalido(Exception):
    """O ajuste foi informado pela metade. Não se completa por conta própria."""


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()




def _linha_informada(valor: Any) -> int | str | None:
    """'003', 3, 3.0 e 'ANOTAR' — tudo o que a pessoa pode escrever."""
    texto = _texto(valor)
    if not texto:
        return None
    if texto.strip().upper() == ANOTAR:
        return ANOTAR
    try:
        return int(float(texto))
    except ValueError:
        return texto


#: Atividades que a segregação de MS reconhece. Só fazem sentido onde a UF
#: segrega — nas demais a coluna fica vazia.
ATIVIDADES = ("industrial", "comercial", "importados", "prestacional_outras")


def _atividade(valor: Any) -> str:
    """Aceita 'Produção' e 'Industrial' pelo que o time fiscal escreve."""
    texto = _texto(valor).casefold().replace(" ", "_")
    if not texto:
        return ""
    apelidos = {"produção": "industrial", "producao": "industrial",
                "prestacional": "prestacional_outras",
                "outras": "prestacional_outras"}
    texto = apelidos.get(texto, texto)
    if texto not in ATIVIDADES:
        raise AjusteInvalido(
            f"atividade {_texto(valor)!r} não existe — as que valem são "
            + ", ".join(ATIVIDADES)
        )
    return texto


def montar(
    estabelecimento: str,
    linha: Any,
    valor: Any,
    motivo: Any,
    responsavel: Any,
    aprovador: Any,
    atividade: Any = None,
    onde: str = "",
) -> Ajuste | None:
    """Valida um ajuste informado e o devolve pronto, ou levanta.

    Devolve `None` quando nada foi informado — a esmagadora maioria das linhas.
    """
    informada = _linha_informada(linha)
    numero = _numero(valor)
    motivo, responsavel, aprovador = (
        _texto(motivo), _texto(responsavel), _texto(aprovador)
    )
    if informada is None and numero is None and not motivo:
        return None
    atividade = _atividade(atividade)

    if informada is None:
        raise AjusteInvalido(
            "ajuste sem `ajuste_linha` — informe 002, 003, 006, 007 ou ANOTAR"
        )
    if informada != ANOTAR and informada not in LINHA_DO_REGISTRO:
        raise AjusteInvalido(
            f"`ajuste_linha` {informada!r} não existe — as linhas que aceitam "
            "ajuste são 002, 003, 006 e 007, mais ANOTAR para marcar sem lançar"
        )
    if not motivo:
        raise AjusteInvalido("ajuste sem `ajuste_motivo` — diga por que ele existe")
    if not responsavel:
        raise AjusteInvalido("ajuste sem `ajuste_responsavel`")

    if informada == ANOTAR:
        return Ajuste(
            estabelecimento=estabelecimento, linha=None, valor=numero or 0.0,
            motivo=motivo, responsavel=responsavel, aprovador=aprovador,
            atividade=atividade, onde=onde,
        )

    if numero is None:
        raise AjusteInvalido(
            f"ajuste na linha {informada:03d} sem `ajuste_valor`"
        )
    if numero < 0:
        raise AjusteInvalido(
            f"`ajuste_valor` {numero} é negativo — o valor é sempre positivo, e "
            "quem dá o sentido é a linha: 002 e 003 aumentam o que se deve, "
            "006 e 007 diminuem"
        )
    if not aprovador:
        raise AjusteInvalido(
            "ajuste sem `ajuste_aprovador` — lançamento que muda o Registro "
            "precisa de quem aprovou"
        )
    return Ajuste(
        estabelecimento=estabelecimento, linha=int(informada), valor=numero,
        motivo=motivo, responsavel=responsavel, aprovador=aprovador,
        atividade=atividade, onde=onde,
    )


def da_linha(dados: dict[str, Any], estabelecimento: str, onde: str) -> Ajuste | None:
    """O ajuste informado nas colunas `ajuste_*` de uma linha da BASE TRATADA."""
    return montar(
        estabelecimento=estabelecimento,
        linha=dados.get("ajuste_linha"),
        valor=dados.get("ajuste_valor"),
        motivo=dados.get("ajuste_motivo"),
        responsavel=dados.get("ajuste_responsavel"),
        aprovador=dados.get("ajuste_aprovador"),
        onde=onde,
    )


@dataclass
class Declarados:
    """O que a aba AJUSTES trouxe."""

    parcelas: list[Ajuste] = field(default_factory=list)
    conferencia: dict[str, Conferencia] = field(default_factory=dict)
    erros: list[str] = field(default_factory=list)

    def __bool__(self) -> bool:
        return bool(self.parcelas or self.conferencia or self.erros)


def ler_aba(caminho) -> Declarados:
    """Lê a aba AJUSTES de um arquivo devolvido. Sem a aba, devolve vazio.

    A aba tem dois blocos, e cada um é achado pelo seu título — não pela linha
    em que está, porque quem preenche insere e apaga linhas.
    """
    from pathlib import Path

    caminho = Path(caminho)
    if caminho.suffix.lower() not in (".xlsx", ".xlsm"):
        return Declarados()

    import openpyxl

    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception:                                       # noqa: BLE001
        return Declarados()
    try:
        if ABA not in wb.sheetnames:
            return Declarados()
        linhas = [list(linha) for linha in wb[ABA].iter_rows(values_only=True)]
    finally:
        wb.close()

    return _interpretar(linhas)


def _interpretar(linhas: list[list[Any]]) -> Declarados:
    declarados = Declarados()
    bloco = ""
    for n, bruta in enumerate(linhas, start=1):
        celulas = list(bruta) + [None] * (7 - len(bruta))
        primeira = _texto(celulas[0])
        if primeira.upper() == TITULO_PARCELAS:
            bloco = TITULO_PARCELAS
            continue
        if primeira.upper() == TITULO_CONFERENCIA:
            bloco = TITULO_CONFERENCIA
            continue
        if not primeira or primeira == "estabelecimento":
            continue

        if bloco == TITULO_PARCELAS:
            try:
                ajuste = montar(
                    estabelecimento=primeira, atividade=celulas[1],
                    linha=celulas[2], valor=celulas[3], motivo=celulas[4],
                    responsavel=celulas[5], aprovador=celulas[6],
                    onde=f"aba {ABA}, linha {n}",
                )
            except AjusteInvalido as erro:
                declarados.erros.append(f"aba {ABA}, linha {n}: {erro}")
                continue
            if ajuste is not None:
                declarados.parcelas.append(ajuste)
        elif bloco == TITULO_CONFERENCIA:
            por = _texto(celulas[1])
            if not por:
                continue                       # ainda não conferido — é o normal
            declarados.conferencia[primeira] = Conferencia(
                estabelecimento=primeira, por=por,
                em=_texto(celulas[2]), observacao=_texto(celulas[3]),
            )
    return declarados
