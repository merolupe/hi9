"""As abas de conferência da apuração.

Três visões da mesma competência, cada uma respondendo a uma pergunta diferente:

`APURAÇÃO EFETIVA`   por que cada crédito foi estornado ou apropriado — CFOP,
                     alíquota e produto, com a conta à vista e o CHECK fechando
                     linha a linha.

`REGISTRO`           o espelho do Registro de Apuração do ICMS, um bloco por
                     estabelecimento e um totalizador no fim — que é o que o
                     PDF do ERP, emitido filial a filial, não mostra.

`TRANSFERÊNCIAS`     o que precisa ser transferido para a centralizadora depois
                     que a competência fechar.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .apuracao import Apuracao, LinhaApurada
from .nucleo import atividade as ativ
from .nucleo import centralizacao as centr
from .nucleo import registro as reg

TITULO = Font(bold=True, color="FFFFFF")
FUNDO = PatternFill("solid", fgColor="1F3864")
FUNDO_CLARO = PatternFill("solid", fgColor="D9E2F3")
FUNDO_ATENCAO = PatternFill("solid", fgColor="FFF2CC")
MOEDA = "#,##0.00"
PERCENTUAL = "0.00%"
VERMELHO = Font(bold=True, color="B00020")
VERDE = Font(bold=True, color="1E7B34")

ROTULO_SEM_ATIVIDADE = "(não segregada)"

#: Como a categoria da operação aparece na conferência. O nome interno é
#: `frete_transferencia`; quem lê a planilha lê "Frete de Transferência".
#: A BASE TRATADA continua com o nome interno de propósito — é por ele que a
#: ferramenta relê o próprio arquivo.
ROTULO_DA_CATEGORIA = {
    "ativo_imobilizado": "Ativo Imobilizado",
    "ciap": "CIAP",
    "complemento_icms": "Complemento de ICMS",
    "complemento_preco": "Complemento de Preço",
    "devolucao_compra": "Devolução de Compra",
    "devolucao_uso_consumo": "Devolução de Uso e Consumo",
    "devolucao_venda": "Devolução de Venda",
    "embalagem": "Embalagem",
    "frete_compra": "Frete de Compra",
    "frete_compra_revenda": "Frete de Compra para Revenda",
    "frete_transferencia": "Frete de Transferência",
    "frete_venda": "Frete de Venda",
    "industrializacao_terceiros": "Industrialização por Terceiros",
    "insumo_energetico": "Insumo Energético",
    "materia_prima": "Matéria-Prima",
    "produto_acabado": "Produto Acabado",
    "produto_quimico": "Produto Químico",
    "quebra": "Quebra",
    "retorno_industrializacao": "Retorno de Industrialização",
    "revenda": "Revenda",
    "uso_consumo": "Uso e Consumo",
}

#: Palavras que ficam em minúscula no meio do rótulo gerado automaticamente.
LIGACOES = {"de", "da", "do", "das", "dos", "e", "por", "para", "em", "a"}


def rotulo_da_categoria(categoria: str) -> str:
    """O nome de exibição de uma categoria.

    Categoria nova que ainda não está no mapa não fica com cara de código: o
    nome interno é quebrado nos underscores e capitalizado, com as ligações em
    minúscula. Cadastrar no mapa continua sendo melhor, porque acento e
    hífen só saem certos lá.
    """
    nome = str(categoria or "").strip()
    if not nome:
        return ""
    if nome in ROTULO_DA_CATEGORIA:
        return ROTULO_DA_CATEGORIA[nome]
    palavras = nome.split("_")
    return " ".join(
        p if i and p in LIGACOES else p.capitalize()
        for i, p in enumerate(palavras)
    )

#: Como a atividade aparece na conferência — os nomes que o time fiscal usa.
ROTULO_DA_ATIVIDADE = {
    ativ.INDUSTRIAL: "Produção",
    ativ.COMERCIAL: "Comercial",
    ativ.IMPORTADOS: "Importados",
    ativ.PRESTACIONAL: "Prestacional / Outras",
}


# --------------------------------------------------------------------------
# Utilitários de escrita
# --------------------------------------------------------------------------

def _larguras(aba, larguras: list[int]) -> None:
    for i, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(i)].width = largura


def _titulo(aba, texto: str, ate: int, *, fundo=FUNDO, fonte=TITULO) -> int:
    linha = aba.max_row + 1
    aba.cell(row=linha, column=1, value=texto)
    for coluna in range(1, ate + 1):
        celula = aba.cell(row=linha, column=coluna)
        celula.fill, celula.font = fundo, fonte
    return linha


def _cabecalho(aba, rotulos: list[str]) -> int:
    linha = aba.max_row + 1
    for i, rotulo in enumerate(rotulos, start=1):
        celula = aba.cell(row=linha, column=i, value=rotulo)
        celula.font, celula.fill = TITULO, FUNDO
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    return linha


def _moeda(aba, linha: int, colunas: range) -> None:
    for coluna in colunas:
        aba.cell(row=linha, column=coluna).number_format = MOEDA


def _vazio(aba) -> None:
    # `append([])` não avança a linha no openpyxl; `append([None])` avança e
    # não escreve nada. É o que dá o respiro entre os blocos.
    aba.append([None])


# --------------------------------------------------------------------------
# APURAÇÃO EFETIVA
# --------------------------------------------------------------------------
#
# A conferência que o time fiscal montava à mão. Duas decisões a governam:
#
# **Agrega no nível do produto.** A apuração manual é uma tabela dinâmica, e
# tabela dinâmica soma: "ÁCIDO FOSFÓRICO RAFINADO" aparece uma vez com o total
# do mês, não uma vez por nota. Listar linha a linha aqui não acrescenta nada
# que a BASE TRATADA já não dê, e enterra a conferência em milhares de linhas.
#
# **Agrupa pela chave da regra do regime, não sempre pela alíquota.** Em MS o
# estorno é uma fração da alíquota; em SP é o excedente da carga efetiva sobre
# a carga de saída. Agrupar pela grandeza errada esconde justamente o que se
# quer conferir.

COLUNAS_EFETIVA = [
    "CFOP", "Descrição", "Carga efetiva", "Produto",
    "Vlr. contábil", "BC ICMS", "Vlr. ICMS", "Operação",
    "% do crédito estornado", "ICMS a estornar", "ICMS a apropriar",
]

#: Colunas por letra, para montar as fórmulas.
COL_CONTABIL, COL_BASE, COL_ICMS = 5, 6, 7
COL_PERCENTUAL, COL_ESTORNAR, COL_APROPRIAR = 9, 10, 11
LETRA = {COL_CONTABIL: "E", COL_BASE: "F", COL_ICMS: "G", COL_ESTORNAR: "J"}

#: Fórmulas de estorno, como aparecem em `regimes.yaml`.
PROPORCIONAL = "proporcional_parcela_nao_tributada"
EXCEDENTE = "excedente_sobre_carga_saida"


def _rotulo_atividade(apurada: LinhaApurada) -> str:
    if not apurada.atividade:
        # A UF não segrega por atividade: mostra a categoria da equalização,
        # que é o corte que faz sentido ali.
        return (
            rotulo_da_categoria(apurada.tratada.classificacao.categoria)
            or ROTULO_SEM_ATIVIDADE
        )
    return ROTULO_DA_ATIVIDADE.get(apurada.atividade, apurada.atividade)


class Somas:
    """Acumulador de um grupo da conferência."""

    __slots__ = ("contabil", "base", "icms", "estornar", "apropriar", "documentos")

    def __init__(self) -> None:
        self.contabil = self.base = self.icms = 0.0
        self.estornar = self.apropriar = 0.0
        self.documentos = 0

    def somar(self, apurada: LinhaApurada) -> None:
        dados = apurada.tratada.origem.dados
        self.contabil += _numero(dados.get("valor_contabil"))
        self.base += _numero(dados.get("base_icms"))
        self.icms += apurada.resultado.credito_bruto
        self.estornar += apurada.credito_a_estornar
        self.apropriar += apurada.credito_a_apropriar
        self.documentos += 1

    def absorver(self, outra: "Somas") -> None:
        self.contabil += outra.contabil
        self.base += outra.base
        self.icms += outra.icms
        self.estornar += outra.estornar
        self.apropriar += outra.apropriar
        self.documentos += outra.documentos

    @property
    def percentual(self) -> float | None:
        """Quanto do crédito a regra mandou estornar, de fato."""
        return (self.estornar / self.icms) if self.icms else None

    @property
    def fecha(self) -> bool:
        """a estornar + a apropriar = crédito. Identidade do motor.

        Não vira coluna na planilha: quem garante isso é o teste de regressão,
        não uma célula que o leitor precisa conferir. Aqui serve só para pintar
        de vermelho a linha que não fechar.
        """
        return abs(self.estornar + self.apropriar - self.icms) < 0.005


def _numero(valor) -> float:
    try:
        return float(valor or 0.0)
    except (TypeError, ValueError):
        return 0.0


def aba_apuracao_efetiva(wb, apuracao: Apuracao, params) -> None:
    """Conferência do crédito, do estorno e da apropriação, por CFOP e produto."""
    aba = wb.create_sheet("APURAÇÃO EFETIVA")
    _larguras(aba, [10, 32, 13, 46, 16, 16, 15, 26, 14, 16, 16])

    aba.append(["APURAÇÃO EFETIVA — crédito, estorno e apropriação por CFOP e produto"])
    aba.cell(row=1, column=1).font = Font(bold=True, size=14)
    aba.append([
        "Um bloco por estabelecimento, agregado como na apuração manual: uma "
        "linha por produto, não por documento. `% do crédito estornado` é "
        "`ICMS a estornar ÷ Vlr. ICMS`. Linha em vermelho é erro de motor: o "
        "estorno mais a apropriação não fecham o crédito."
    ])

    for filial in sorted(
        apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)
    ):
        _bloco_efetiva(aba, filial)

    aba.freeze_panes = "A3"


def _bloco_efetiva(aba, filial) -> None:
    # Respiro entre estabelecimentos — os blocos ficavam colados.
    _vazio(aba)
    _vazio(aba)
    _titulo(
        aba,
        f"{filial.estabelecimento}  —  {filial.uf}  —  regime {filial.regime}",
        len(COLUNAS_EFETIVA),
    )
    entradas = [a for a in filial.apuradas if a.resultado.credito_bruto]
    if not entradas:
        aba.append(["", "Sem crédito de entrada nesta competência."])
        _blocos_de_fechamento(aba, filial)
        return

    _cabecalho(aba, COLUNAS_EFETIVA)

    # CFOP → CARGA EFETIVA EQUALIZADA → produto × operação, somando em cada
    # nível. A carga é a chave nos dois regimes: é a grandeza que o documento
    # traz depois da equalização, e é por ela que a conferência manual olha.
    arvore: dict = {}
    for apurada in entradas:
        cfop = apurada.tratada.origem.cfop_int
        carga = apurada.tratada.carga.carga
        produto = (_nome_do_produto(apurada), _rotulo_atividade(apurada))
        por_cfop = arvore.setdefault(cfop, {"descricao": "", "cargas": {}})
        por_cfop["descricao"] = por_cfop["descricao"] or _descricao_cfop(apurada)
        por_carga = por_cfop["cargas"].setdefault(carga, {})
        por_carga.setdefault(produto, Somas()).somar(apurada)

    geral, linhas_de_cfop = Somas(), []
    for cfop in sorted(arvore, key=lambda c: (c is None, c or 0)):
        ramo = arvore[cfop]
        do_cfop = _agregar(ramo["cargas"].values())
        linha_cfop = _linha_de_grupo(
            aba, do_cfop, nivel=0,
            chave=cfop if cfop is not None else "(sem CFOP)",
            descricao=ramo["descricao"],
        )
        linhas_de_carga = []
        for carga in sorted(ramo["cargas"], key=lambda v: (v is None, v or 0)):
            produtos = ramo["cargas"][carga]
            linha_carga = _linha_de_grupo(
                aba, _agregar([produtos]), nivel=1, valor_da_chave=carga
            )
            folhas = [
                _linha_de_produto(aba, produtos[chave], carga, *chave)
                for chave in sorted(produtos, key=lambda p: (p[0].casefold(), p[1]))
            ]
            _somatorio(aba, linha_carga, folhas)
            linhas_de_carga.append(linha_carga)
        _somatorio(aba, linha_cfop, linhas_de_carga)
        linhas_de_cfop.append(linha_cfop)
        geral.absorver(do_cfop)

    linha_total = _linha_de_grupo(aba, geral, nivel=0, chave="TOTAL", negrito=True)
    _somatorio(aba, linha_total, linhas_de_cfop)
    _blocos_de_fechamento(aba, filial)


def _agregar(grupos, inicial: Somas | None = None) -> Somas:
    total = inicial or Somas()
    for grupo in grupos:
        for parcial in grupo.values():
            total.absorver(parcial)
    return total


def _nome_do_produto(apurada: LinhaApurada) -> str:
    return str(apurada.tratada.origem.dados.get("produto_descricao") or "")


def _descricao_cfop(apurada: LinhaApurada) -> str:
    return str(apurada.tratada.origem.dados.get("cfop_descricao") or "").strip()


def _percentual(valor: float | None) -> str:
    return "" if valor is None else f"{valor:g}%"


def _somatorio(aba, linha: int, filhas: list[int]) -> None:
    """Troca os valores colados do grupo por SOMA das linhas que o compõem.

    O motor continua sendo quem calcula; a planilha passa a mostrar de onde
    cada total veio, e recalcula sozinha se alguém mexer numa linha.
    """
    if not filhas:
        return
    contiguas = filhas == list(range(filhas[0], filhas[-1] + 1))
    for coluna, letra in LETRA.items():
        alvo = (
            f"{letra}{filhas[0]}:{letra}{filhas[-1]}" if contiguas
            else ",".join(f"{letra}{f}" for f in filhas)
        )
        aba.cell(row=linha, column=coluna).value = f"=SUM({alvo})"


def _derivadas(aba, linha: int) -> None:
    """As duas colunas que são identidade, não regra — vão como fórmula.

    % do crédito estornado = a estornar ÷ crédito
    ICMS a apropriar       = crédito − a estornar
    """
    aba.cell(row=linha, column=COL_PERCENTUAL).value = (
        f'=IF(G{linha}=0,"",J{linha}/G{linha})'
    )
    aba.cell(row=linha, column=COL_APROPRIAR).value = f"=G{linha}-J{linha}"


def _formatos_da_linha(aba, linha: int) -> None:
    _moeda(aba, linha, range(COL_CONTABIL, COL_ICMS + 1))
    _moeda(aba, linha, range(COL_ESTORNAR, COL_APROPRIAR + 1))
    aba.cell(row=linha, column=COL_PERCENTUAL).number_format = PERCENTUAL


def _linha_de_grupo(
    aba, somas: Somas, *, nivel: int, chave=None, descricao: str = "",
    valor_da_chave: float | None = None, negrito: bool = False,
) -> int:
    aba.append([
        chave if nivel == 0 else "", descricao,
        _percentual(valor_da_chave) if nivel else "", "",
        somas.contabil, somas.base, somas.icms, "",
        None, somas.estornar, None,
    ])
    linha = aba.max_row
    _derivadas(aba, linha)
    _formatos_da_linha(aba, linha)
    for celula in aba[linha]:
        celula.font = Font(bold=True)
        if nivel == 0 and not negrito:
            celula.fill = FUNDO_CLARO
    if nivel:
        aba.row_dimensions[linha].outlineLevel = 1
    return linha


def _linha_de_produto(
    aba, somas: Somas, valor_da_chave: float | None, produto: str, operacao: str
) -> int:
    aba.append([
        "", "", _percentual(valor_da_chave), produto,
        somas.contabil, somas.base, somas.icms, operacao,
        None, somas.estornar, None,
    ])
    linha = aba.max_row
    _derivadas(aba, linha)
    _formatos_da_linha(aba, linha)
    aba.row_dimensions[linha].outlineLevel = 2
    if not somas.fecha:
        # A identidade quebrou. Sem coluna de CHECK, o aviso é a própria linha.
        for celula in aba[linha]:
            celula.font = VERMELHO
    return linha


def _blocos_de_fechamento(aba, filial) -> None:
    """Créditos e débitos por classificação — o fechamento que a GIA pede.

    A carga efetiva de cada classificação é a MÉDIA PONDERADA pelo valor
    contábil das linhas que entraram nela: onde a classificação tem uma carga
    só, o número sai redondo; onde mistura, o resultado diz onde ela está.
    """
    por_atividade: dict[str, dict[str, float]] = {}
    for a in filial.apuradas:
        chave = _rotulo_atividade(a)
        alvo = por_atividade.setdefault(
            chave, {"bc_credito": 0.0, "credito": 0.0, "estorno": 0.0,
                    "bc_debito": 0.0, "debito": 0.0,
                    "contabil": 0.0, "contabil_x_carga": 0.0,
                    "icms_x_carga": 0.0}
        )
        base = _numero(a.tratada.origem.dados.get("base_icms"))
        if a.resultado.credito_bruto:
            alvo["bc_credito"] += base
            alvo["credito"] += a.resultado.credito_bruto
            alvo["estorno"] += a.credito_a_estornar
            contabil = _numero(a.tratada.origem.dados.get("valor_contabil"))
            carga = a.tratada.carga.carga
            if carga is not None:
                alvo["contabil"] += contabil
                alvo["contabil_x_carga"] += contabil * carga
                alvo["icms_x_carga"] += a.resultado.credito_bruto * carga
        if a.resultado.debito:
            alvo["bc_debito"] += base
            alvo["debito"] += a.resultado.debito

    _vazio(aba)
    _titulo(aba, "CRÉDITOS", 7, fundo=FUNDO_CLARO, fonte=Font(bold=True))
    _cabecalho(aba, ["", "Classificação", "Carga efetiva", "BC ICMS",
                     "VLR ICMS", "ESTORNO ICMS", "A APROPRIAR"])
    primeira, soma = aba.max_row + 1, [0.0, 0.0, 0.0]
    for nome in sorted(por_atividade):
        v = por_atividade[nome]
        if not v["credito"]:
            continue
        aba.append(["", nome, _carga_media(v), v["bc_credito"], v["credito"],
                    v["estorno"]])
        linha = aba.max_row
        aba.cell(row=linha, column=7).value = f"=E{linha}-F{linha}"
        _moeda(aba, linha, range(4, 8))
        aba.cell(row=linha, column=3).number_format = PERCENTUAL
        soma = [soma[0] + v["bc_credito"], soma[1] + v["credito"],
                soma[2] + v["estorno"]]
    aba.append(["", "TOTAL", None, soma[0], soma[1], soma[2]])
    linha = aba.max_row
    if linha > primeira:
        for coluna, letra in ((4, "D"), (5, "E"), (6, "F")):
            aba.cell(row=linha, column=coluna).value = (
                f"=SUM({letra}{primeira}:{letra}{linha - 1})"
            )
    aba.cell(row=linha, column=7).value = f"=E{linha}-F{linha}"
    _moeda(aba, linha, range(4, 8))
    for celula in aba[linha]:
        celula.font = Font(bold=True)

    _vazio(aba)
    _titulo(aba, "DÉBITOS", 7, fundo=FUNDO_CLARO, fonte=Font(bold=True))
    _cabecalho(aba, ["", "Classificação", "", "BC ICMS", "VLR ICMS"])
    primeira, soma_debito = aba.max_row + 1, [0.0, 0.0]
    for nome in sorted(por_atividade):
        v = por_atividade[nome]
        if not v["debito"]:
            continue
        aba.append(["", nome, None, v["bc_debito"], v["debito"]])
        _moeda(aba, aba.max_row, range(4, 6))
        soma_debito = [soma_debito[0] + v["bc_debito"], soma_debito[1] + v["debito"]]
    aba.append(["", "TOTAL", None, soma_debito[0], soma_debito[1]])
    linha = aba.max_row
    if linha > primeira:
        for coluna, letra in ((4, "D"), (5, "E")):
            aba.cell(row=linha, column=coluna).value = (
                f"=SUM({letra}{primeira}:{letra}{linha - 1})"
            )
    _moeda(aba, linha, range(4, 6))
    for celula in aba[linha]:
        celula.font = Font(bold=True)

    if filial.beneficio:
        _vazio(aba)
        _titulo(aba, "BENEFÍCIO FISCAL", 7, fundo=FUNDO_CLARO, fonte=Font(bold=True))
        for passo in filial.beneficio.memoria:
            aba.append(["", passo])


def _carga_media(valores: dict[str, float]) -> float | None:
    """Carga efetiva do grupo, ponderada pelo valor contábil.

    Quando o grupo inteiro não tem valor contábil, a ponderação cai para o
    ICMS. É o caso do complemento de ICMS, que só traz base e imposto — a
    carga dele existe (equalizada como qualquer outra), e ponderar pelo
    contábil deixava a célula vazia como se o dado não existisse.
    """
    if valores["contabil"]:
        return valores["contabil_x_carga"] / valores["contabil"] / 100.0
    if valores["credito"]:
        return valores["icms_x_carga"] / valores["credito"] / 100.0
    return None


# --------------------------------------------------------------------------
# REGISTRO
# --------------------------------------------------------------------------

COLUNAS_REGISTRO = [12, 40, 18, 18, 18, 18, 18]

CABECALHO_VALORES = [
    "Valores Contábeis", "Base de Cálculo", "Imposto {verbo}",
    "Isentas / N. Trib.", "Outras",
]


def aba_registro(wb, apuracao: Apuracao, params, ajustes=None) -> list:
    """Espelho do Registro de Apuração — um bloco por filial e o totalizador."""
    aba = wb.create_sheet("REGISTRO")
    _larguras(aba, COLUNAS_REGISTRO)

    aba.append(["REGISTRO DE APURAÇÃO DO ICMS"])
    aba.cell(row=1, column=1).font = Font(bold=True, size=14)
    aba.append([
        "Espelho do livro por estabelecimento. As linhas do resumo marcadas "
        "AGUARDA AJUSTE dependem de lançamento aprovado que não nasce do Livro Fiscal."
    ])

    registros = reg.montar(apuracao, params, ajustes)
    for registro in registros:
        _bloco_registro(aba, registro)
    if len(registros) > 1:
        _bloco_registro(aba, reg.totalizador(registros, apuracao.base.competencia))

    aba.freeze_panes = "A3"
    return registros


def _bloco_registro(aba, registro: reg.Registro) -> None:
    _vazio(aba)
    _vazio(aba)
    _titulo(aba, registro.estabelecimento, 7)
    if registro.gerencial:
        aba.append([
            "", "Soma dos estabelecimentos. Não é documento fiscal: consolida "
            "UFs com contas gráficas distintas."
        ])
    else:
        aba.append(["FIRMA", registro.estabelecimento, "UF", registro.uf])
        aba.append([
            "CNPJ", registro.cnpj or "(não cadastrado)",
            "INSCRIÇÃO ESTADUAL", registro.inscricao_estadual or "(não cadastrado)",
        ])
        aba.append(["PERÍODO", registro.competencia])

    _bloco_de_valores(aba, registro.entradas, "Creditado")
    _bloco_de_valores(aba, registro.saidas, "Debitado")
    _bloco_de_resumo(aba, registro)


def _bloco_de_valores(aba, bloco: reg.Bloco, verbo: str) -> None:
    _vazio(aba)
    _titulo(aba, bloco.lado, 7, fundo=FUNDO_CLARO, fonte=Font(bold=True))
    _cabecalho(
        aba,
        ["CFOP", "Codificação Contábil-Fiscal"]
        + [c.format(verbo=verbo) for c in CABECALHO_VALORES],
    )
    for linha in bloco.linhas:
        aba.append([linha.cfop, linha.descricao, *linha.valores.as_tuple()])
        _moeda(aba, aba.max_row, range(3, 8))

    subtotais = []
    for grupo in bloco.grupos():
        valores = bloco.subtotal(grupo)
        aba.append(["", f"Subtotal {grupo}", *valores.as_tuple()])
        subtotais.append(aba.max_row)
        _moeda(aba, aba.max_row, range(3, 8))
        for celula in aba[aba.max_row]:
            celula.fill = FUNDO_CLARO

    aba.append(["", "TOTAL", *bloco.total.as_tuple()])
    linha = aba.max_row
    # O TOTAL soma os subtotais na própria planilha: quem confere vê de onde
    # ele veio, e o arquivo recalcula sozinho se alguém mexer numa linha.
    if subtotais:
        for coluna in range(3, 8):
            letra = get_column_letter(coluna)
            aba.cell(row=linha, column=coluna).value = (
                f"=SUM({','.join(f'{letra}{n}' for n in subtotais)})"
            )
    _moeda(aba, linha, range(3, 8))
    for celula in aba[linha]:
        celula.font = Font(bold=True)


#: Linhas que consolidam as anteriores — vão na coluna "Somas".
LINHAS_DE_SOMA = {4, 8, 10, 11, 13, 14}

#: As três somas do resumo que são aritmética pura das linhas acima, e por isso
#: podem ir como fórmula. A coluna de cada parcela vai junto: D é "Valores",
#: onde a linha comum grava; E é "Somas", onde a linha consolidada grava.
#:
#: 011, 013 e 014 ficam de fora de propósito: dependem do SINAL do resultado —
#: quem apura devedor preenche 011 e 013 e zera 014, quem apura credor faz o
#: contrário. Isso é decisão do motor, não conta de planilha.
SOMAS_DO_RESUMO = {
    4: [("D", 1), ("D", 2), ("D", 3)],
    8: [("D", 5), ("D", 6), ("D", 7)],
    10: [("E", 8), ("D", 9)],
}


def _somar_no_resumo(aba, codigo: int, onde: dict[int, int]) -> None:
    """Troca o valor colado da linha consolidada pela soma das parcelas."""
    parcelas = SOMAS_DO_RESUMO.get(codigo)
    if not parcelas or not all(c in onde for _, c in parcelas):
        return
    termos = "+".join(f"{letra}{onde[c]}" for letra, c in parcelas)
    aba.cell(row=onde[codigo], column=5).value = f"={termos}"


def _bloco_de_resumo(aba, registro: reg.Registro) -> None:
    _vazio(aba)
    _titulo(aba, "RESUMO DA APURAÇÃO DO IMPOSTO", 7,
            fundo=FUNDO_CLARO, fonte=Font(bold=True))
    _cabecalho(aba, ["Código", "Descrição", "Coluna Auxiliar", "Valores", "Somas",
                     "", "Situação"])
    if registro.gerencial:
        aba.append([
            "", "As linhas 011 a 014 são a soma do resultado de cada "
            "estabelecimento, não o recálculo sobre os totais: crédito de uma UF "
            "não abate débito de outra."
        ])
        aba.cell(row=aba.max_row, column=2).fill = FUNDO_ATENCAO

    onde: dict[int, int] = {}
    for item in registro.resumo:
        soma = item.codigo in LINHAS_DE_SOMA
        aba.append([
            f"{item.codigo:03d}", item.rotulo, None,
            None if soma else item.valor,
            item.valor if soma else None, "",
            "AGUARDA AJUSTE" if item.aguarda_ajuste else "",
        ])
        linha = onde[item.codigo] = aba.max_row
        _somar_no_resumo(aba, item.codigo, onde)
        _moeda(aba, linha, range(3, 6))
        if soma:
            for celula in aba[linha]:
                celula.font = Font(bold=True)
        if item.aguarda_ajuste:
            aba.cell(row=linha, column=7).fill = FUNDO_ATENCAO
        if item.codigo == 13:
            aba.cell(row=linha, column=5).font = VERMELHO if item.valor else VERDE

        for descricao, valor in item.discriminacao:
            aba.append(["", f"    {descricao}", valor])
            aba.cell(row=aba.max_row, column=3).number_format = MOEDA
            aba.row_dimensions[aba.max_row].outlineLevel = 1


# --------------------------------------------------------------------------
# TRANSFERÊNCIAS
# --------------------------------------------------------------------------

def aba_transferencias(wb, apuracao: Apuracao) -> None:
    """O que transferir para a centralizadora depois de fechar a competência."""
    aba = wb.create_sheet("TRANSFERÊNCIAS")
    _larguras(aba, [34, 34, 18, 18, 18, 30, 22, 16])

    aba.append(["TRANSFERÊNCIAS DE SALDO A EMITIR"])
    aba.cell(row=1, column=1).font = Font(bold=True, size=14)
    aba.append([
        "A transferência é consequência da apuração: o documento que a formaliza "
        "só pode ser emitido depois do encerramento e vai escriturado na "
        "competência seguinte. Esta aba é a instrução, não uma conferência."
    ])

    if not apuracao.centralizacao:
        _vazio(aba)
        aba.append(["", "Nenhuma UF com apuração centralizada parametrizada."])
        return

    for grupo in apuracao.centralizacao:
        _vazio(aba)
        _titulo(aba, f"{grupo.uf} — centraliza em {grupo.centralizadora}", 8)
        if not grupo.homologado:
            aba.append([
                "", "REGRA NÃO HOMOLOGADA — falta a Gerência Fiscal/Tributária "
                "confirmar o que se transfere e por qual documento."
            ])
            aba.cell(row=aba.max_row, column=2).fill = FUNDO_ATENCAO

        _cabecalho(aba, [
            "Origem", "Destino", "Saldo do estabelecimento", "A transferir",
            "Saldo residual", "Mecanismo", "CFOP sugerido", "Confere",
        ])
        for t in grupo.transferencias:
            aba.append([
                t.origem, t.destino, t.saldo_individual, t.valor_transferido,
                t.saldo_residual,
                centr.MECANISMOS.get(t.mecanismo, t.mecanismo),
                ", ".join(str(c) for c in t.cfop_sugerido) if t.cfop_sugerido else "",
                "OK" if t.confere else "DIVERGE",
            ])
            _moeda(aba, aba.max_row, range(3, 6))

        _vazio(aba)
        for rotulo, valor in (
            ("Saldo próprio da centralizadora", grupo.saldo_proprio),
            ("Recebido dos centralizados", grupo.total_recebido),
            ("Saldo final do grupo", grupo.saldo_final),
        ):
            aba.append(["", rotulo, valor])
            aba.cell(row=aba.max_row, column=3).number_format = MOEDA
        for celula in aba[aba.max_row]:
            celula.font = Font(bold=True)

        _vazio(aba)
        aba.append(["", "O que emitir:"])
        aba.cell(row=aba.max_row, column=2).font = Font(bold=True)
        for instrucao in grupo.instrucoes:
            aba.append(["", instrucao])
        if not grupo.instrucoes:
            aba.append(["", "Nada a transferir nesta competência."])
