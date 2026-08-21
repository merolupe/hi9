"""Camada 1 e 2 — ingestão e normalização do Livro Fiscal.

Lê o arquivo extraído do Sankhya, confere o layout contra o dicionário de
dados e devolve linhas normalizadas. Não interpreta regra tributária.
"""
from __future__ import annotations

import datetime as dt
import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

# Nome da coluna no arquivo -> nome do campo normalizado.
# A ordem não importa; a presença, sim. Ver docs/apurabot/03-dicionario-livro-fiscal.md
COLUNAS = {
    "Report": "report",
    "Status": "status_documento",
    "Data Inclusão/Alteração": "data_inclusao",
    "Nro único Nota": "nro_unico",
    "Diferença ICMS": "diferenca_icms",
    "Sequência": "sequencia",
    "Empresa": "empresa_codigo",
    "Nome Fantasia (Empresa)": "estabelecimento",
    "Dt. do documento": "data_documento",
    "Dt. do movimento": "data_movimento",
    "Empresa/Parceiro": "parceiro_codigo",
    "Descrição Parceiro/Empresa": "parceiro",
    "Nro. da nota": "numero_nota",
    "CFOP": "cfop",
    "Descrição da CFOP": "cfop_descricao",
    "Cód. de tributação": "cst",
    "Produto": "produto",
    "Descrição (Produto)": "produto_descricao",
    "Vlr. contábil": "valor_contabil",
    "Base do ICMS": "base_icms",
    "Alíquota ICMS": "aliquota_icms",
    "Vlr. do ICMS": "valor_icms",
    "Carga": "carga_bruta_origem",
    "Carga efetiva": "carga_efetiva_origem",
    "Isentas de ICMS": "isentas_icms",
    "Outras ICMS": "outras_icms",
    "Vlr. do IPI": "valor_ipi",
    "UF de Origem": "uf_origem",
    "UF de Destino": "uf_destino",
    "Série da nota": "serie",
    "Destino": "destino",
    "Dt. Cancelamento": "data_cancelamento",
    "Espécie do documento": "especie",
    "Tipo de ICMS": "tipo_icms",
    "Base retenção": "base_retencao",
    "ICMS retenção": "icms_retencao",
    "Tipo de IPI": "tipo_ipi",
    "Base do IPI": "base_ipi",
    "Alíquota de IPI": "aliquota_ipi",
    "Isentas de IPI": "isentas_ipi",
    "Outras IPI": "outras_ipi",
    "Entrada/Saída": "entrada_saida",
    "Modelo do Documento": "modelo",
    "Chave NF-e": "chave_nfe",
    "Chave CT-e": "chave_cte",
    "Chave CT-e de Referência": "chave_cte_referencia",
    "Cód. Cid. Inicio CT-e": "cidade_origem_codigo",
    "Nome (Cidade de Origem)": "cidade_origem",
    "Cód. Cid. Fim CT-e": "cidade_destino_codigo",
    "Nome (Cidade de Destino)": "cidade_destino",
    "Vlr. ICMS Complemento": "icms_complemento",
    "Nome Fantasia (Empresa Origem)": "estabelecimento_origem",
}

# Sem estas o motor não roda. As demais podem faltar sem quebrar a apuração.
ESSENCIAIS = (
    "Nro único Nota", "Nome Fantasia (Empresa)", "Dt. do movimento", "CFOP",
    "Cód. de tributação", "Vlr. contábil", "Base do ICMS", "Alíquota ICMS",
    "Vlr. do ICMS", "UF de Origem", "UF de Destino", "Entrada/Saída",
    "Espécie do documento", "Produto",
)


class LayoutInvalido(Exception):
    """O arquivo não tem o layout esperado do Livro Fiscal."""


@dataclass
class LinhaLivro:
    """Uma linha do Livro Fiscal, normalizada. Espelha o arquivo, sem regra."""

    linha_origem: int
    arquivo_origem: str
    dados: dict[str, Any] = field(repr=False, default_factory=dict)

    def __getattr__(self, nome: str) -> Any:
        try:
            return self.__dict__["dados"][nome]
        except KeyError:
            raise AttributeError(nome) from None

    # -- conveniências usadas pelo núcleo ------------------------------

    @property
    def cancelado(self) -> bool:
        return self.dados.get("data_cancelamento") is not None

    @property
    def interestadual(self) -> bool:
        return self.dados.get("uf_origem") != self.dados.get("uf_destino")

    @property
    def cfop_int(self) -> int | None:
        cfop = self.dados.get("cfop")
        return int(cfop) if cfop is not None else None

    @property
    def cst_codigo(self) -> str:
        """Só o código do CST: '20-Com redução de base' -> '20'."""
        return str(self.dados.get("cst") or "").strip()[:2]

    @property
    def produto_codigo(self) -> str:
        produto = self.dados.get("produto")
        return "" if produto is None else str(int(produto))


@dataclass
class Livro:
    """O Livro Fiscal inteiro, mais a procedência do arquivo."""

    linhas: list[LinhaLivro]
    arquivo: Path
    sha256: str
    colunas_ausentes: list[str]

    def __iter__(self) -> Iterator[LinhaLivro]:
        return iter(self.linhas)

    def __len__(self) -> int:
        return len(self.linhas)

    @property
    def competencias(self) -> set[str]:
        """Competências encontradas, no formato AAAA-MM."""
        return {
            d.strftime("%Y-%m")
            for linha in self.linhas
            if isinstance(d := linha.dados.get("data_movimento"), dt.date)
        }


def _sha256(caminho: Path) -> str:
    h = hashlib.sha256()
    with caminho.open("rb") as f:
        for bloco in iter(lambda: f.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def _linhas_xls(caminho: Path, aba: str | None) -> tuple[list[str], list[list[Any]]]:
    import xlrd

    livro = xlrd.open_workbook(caminho, on_demand=True)
    candidatas = [
        (nome, _cabecalho_xls(livro, nome), livro.sheet_by_name(nome).nrows)
        for nome in livro.sheet_names()
    ]
    aba = aba or _escolher_aba(candidatas)
    s = livro.sheet_by_name(aba)
    cabecalho = [str(s.cell(0, c).value).strip() for c in range(s.ncols)]

    def valor(celula):
        if celula.ctype in (0, 5, 6):          # vazia, erro, em branco
            return None
        if celula.ctype == 3:                  # data
            try:
                return xlrd.xldate.xldate_as_datetime(celula.value, livro.datemode).date()
            except Exception:
                return None
        if celula.ctype == 4:
            return bool(celula.value)
        if isinstance(celula.value, str):
            return celula.value.strip() or None
        return celula.value

    corpo = [[valor(s.cell(r, c)) for c in range(s.ncols)] for r in range(1, s.nrows)]
    return cabecalho, corpo


def _cabecalho_xls(livro, nome_aba: str) -> list[str]:
    s = livro.sheet_by_name(nome_aba)
    if s.nrows == 0:
        return []
    return [str(s.cell(0, c).value).strip() for c in range(s.ncols)]


def _linhas_xlsx(caminho: Path, aba: str | None) -> tuple[list[str], list[list[Any]]]:
    import openpyxl

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    def cabecalho_de(nome: str) -> list[str]:
        for linha in wb[nome].iter_rows(min_row=1, max_row=1, values_only=True):
            return [str(c).strip() if c is not None else "" for c in linha]
        return []

    candidatas = [
        (nome, cabecalho_de(nome), wb[nome].max_row or 0) for nome in wb.sheetnames
    ]
    nome = aba or _escolher_aba(candidatas)
    aba = wb[nome]
    it = aba.iter_rows(values_only=True)
    cabecalho = [str(c).strip() if c is not None else "" for c in next(it)]

    def valor(v):
        if isinstance(v, str):
            return v.strip() or None
        if isinstance(v, dt.datetime):
            return v.date()
        return v

    corpo = [[valor(v) for v in linha] for linha in it]
    wb.close()
    return cabecalho, corpo


def _escolher_aba(candidatas: list[tuple[str, list[str], int]]) -> str:
    """Escolhe a aba do Livro Fiscal entre as abas do arquivo.

    Em produção o arquivo tem uma aba só e a escolha é trivial. O critério
    existe para o arquivo de análise, que carrega o Livro Fiscal ao lado de
    recortes dele com o mesmo cabeçalho.

    Critério, nesta ordem: (1) mais colunas essenciais no cabeçalho;
    (2) nome parecido com "livro fiscal"; (3) mais linhas — um recorte nunca
    é maior que o livro de onde saiu.
    """
    completas = [
        (nome, cab, nlinhas)
        for nome, cab, nlinhas in candidatas
        if sum(1 for c in ESSENCIAIS if c in set(cab)) == len(ESSENCIAIS)
    ]
    if not completas:
        melhor_nota = max(
            (sum(1 for c in ESSENCIAIS if c in set(cab)) for _, cab, _ in candidatas),
            default=0,
        )
        raise LayoutInvalido(
            "nenhuma aba do arquivo parece ser o Livro Fiscal — a melhor "
            f"candidata tem {melhor_nota} de {len(ESSENCIAIS)} colunas essenciais"
        )

    def prioridade(item: tuple[str, list[str], int]) -> tuple[int, int]:
        nome, _, nlinhas = item
        return (1 if "livro fiscal" in nome.strip().lower() else 0, nlinhas)

    return max(completas, key=prioridade)[0]


def ler_livro_fiscal(caminho: Path | str, aba: str | None = None) -> Livro:
    """Lê o Livro Fiscal (.xlsx ou .xls) e valida o layout.

    Levanta `LayoutInvalido` com mensagem clara se o Sankhya mudar o extrato.
    Passe `aba` para forçar uma aba específica em vez de deixar a escolha
    automática.
    """
    caminho = Path(caminho)
    if not caminho.is_file():
        raise FileNotFoundError(f"Livro Fiscal não encontrado: {caminho}")

    sufixo = caminho.suffix.lower()
    if sufixo == ".xls":
        cabecalho, corpo = _linhas_xls(caminho, aba)
    elif sufixo in (".xlsx", ".xlsm"):
        cabecalho, corpo = _linhas_xlsx(caminho, aba)
    else:
        raise LayoutInvalido(
            f"extensão não suportada: {sufixo!r}. Esperado .xlsx, .xlsm ou .xls"
        )

    presentes = set(cabecalho)
    faltando_essenciais = [c for c in ESSENCIAIS if c not in presentes]
    if faltando_essenciais:
        raise LayoutInvalido(
            "o layout do Livro Fiscal mudou — faltam colunas essenciais:\n  "
            + "\n  ".join(faltando_essenciais)
            + "\n\nConfira o extrato do Sankhya contra "
            "docs/apurabot/03-dicionario-livro-fiscal.md"
        )

    indices = {COLUNAS[c]: i for i, c in enumerate(cabecalho) if c in COLUNAS}
    ausentes = sorted(c for c in COLUNAS if c not in presentes)

    linhas = []
    for n, bruta in enumerate(corpo, start=2):     # linha 1 é o cabeçalho
        dados = {
            campo: (bruta[i] if i < len(bruta) else None)
            for campo, i in indices.items()
        }
        if all(v is None for v in dados.values()):
            continue                               # linha em branco no fim
        linhas.append(
            LinhaLivro(linha_origem=n, arquivo_origem=caminho.name, dados=dados)
        )

    return Livro(
        linhas=linhas,
        arquivo=caminho,
        sha256=_sha256(caminho),
        colunas_ausentes=ausentes,
    )
