"""Escrita da base tratada em .xlsx.

O time fiscal continua recebendo uma planilha — o Python é o motor, não a
interface. Toda linha carrega a rastreabilidade exigida pelo Anexo B do escopo.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import __version__
from .apuracao import Apuracao, apurar
from .conferencia import aba_apuracao_efetiva, aba_registro, aba_transferencias
from .nucleo.atividade import INTERESTADUAL, INTRAESTADUAL
from .nucleo.atividade import ORDEM as ATIVIDADES_EM_ORDEM
from .base_tratada import BaseTratada

#: Procedência — o que a ferramenta escreve, não o Livro.
COLUNAS_DE_PROCEDENCIA = [
    ("arquivo_origem", 22), ("linha_origem", 12), ("competencia", 12),
]

#: Campos do Livro que abrem a aba: são os que se olha primeiro.
COLUNAS_EM_DESTAQUE = [
    ("estabelecimento", 30), ("uf_origem", 10), ("uf_destino", 10),
    ("entrada_saida", 13), ("nro_unico", 12), ("numero_nota", 12),
    ("cfop", 8), ("cfop_descricao", 30), ("cst", 10), ("especie", 9),
    ("produto", 12), ("produto_descricao", 38),
    ("valor_contabil", 15), ("base_icms", 15), ("aliquota_icms", 12),
    ("valor_icms", 15),
]

#: O que o motor concluiu sobre a linha. Na releitura estas colunas são
#: ignoradas e recalculadas: se alguém as editar, a edição não vale.
COLUNAS_CALCULADAS = [
    ("carga_bruta", 13), ("carga_efetiva", 13), ("situacao", 22),
    ("categoria", 26), ("regra_carga", 60), ("regra_classificacao", 60),
    ("alerta", 24), ("pendencia", 60),
]

#: As colunas que o time fiscal preenche. Ver `ajustes.py`.
COLUNAS_DE_AJUSTE = [
    ("ajuste_linha", 14), ("ajuste_valor", 15), ("ajuste_motivo", 46),
    ("ajuste_responsavel", 20), ("ajuste_aprovador", 20),
]


def _demais_campos_do_livro() -> list[tuple[str, int]]:
    """Todo campo do Livro que não está em destaque, na ordem do extrato.

    Eles saem porque a conferência precisa deles — parceiro, datas, série,
    chave, observação, TOP — e porque é o que torna o arquivo devolvido
    autossuficiente: com o Livro inteiro dentro, realimentar é arrastar um
    arquivo só.
    """
    from .ingestao import COLUNAS

    destaque = {nome for nome, _ in COLUNAS_EM_DESTAQUE}
    vistos: set[str] = set()
    colunas = []
    for campo in COLUNAS.values():
        if campo in destaque or campo in vistos:
            continue
        vistos.add(campo)
        colunas.append((campo, max(12, min(30, len(campo) + 4))))
    return colunas


CABECALHO_BASE = (
    COLUNAS_DE_PROCEDENCIA
    + COLUNAS_EM_DESTAQUE
    + COLUNAS_CALCULADAS
    + COLUNAS_DE_AJUSTE
    + _demais_campos_do_livro()
)

#: Onde começam as colunas de ajuste (1-based), para pintá-las de outra cor.
PRIMEIRA_DE_AJUSTE = (
    len(COLUNAS_DE_PROCEDENCIA) + len(COLUNAS_EM_DESTAQUE)
    + len(COLUNAS_CALCULADAS) + 1
)

TITULO = Font(bold=True, color="FFFFFF")
FUNDO = PatternFill("solid", fgColor="1F3864")
MOEDA = "#,##0.00"
PERCENTUAL = "0.00%"


def _escreve_cabecalho(aba, colunas):
    aba.append([nome for nome, _ in colunas])
    for i, (_, largura) in enumerate(colunas, start=1):
        aba.column_dimensions[get_column_letter(i)].width = largura
        celula = aba.cell(row=1, column=i)
        celula.font, celula.fill = TITULO, FUNDO
        celula.alignment = Alignment(vertical="center", wrap_text=False)
    aba.freeze_panes = "A2"


#: Fundo das colunas que o time fiscal preenche — para não se confundirem com
#: as que a ferramenta escreve.
FUNDO_DE_AJUSTE = PatternFill("solid", fgColor="7B3F00")


def _aba_base(wb, base: BaseTratada) -> None:
    aba = wb.create_sheet("BASE TRATADA")
    _escreve_cabecalho(aba, CABECALHO_BASE)
    for i in range(PRIMEIRA_DE_AJUSTE, PRIMEIRA_DE_AJUSTE + len(COLUNAS_DE_AJUSTE)):
        aba.cell(row=1, column=i).fill = FUNDO_DE_AJUSTE

    competencia = base.competencia
    demais = [nome for nome, _ in _demais_campos_do_livro()]
    for t in base.linhas:
        d = t.origem.dados
        aba.append([
            t.origem.arquivo_origem, t.origem.linha_origem, competencia,
            d.get("estabelecimento"), d.get("uf_origem"), d.get("uf_destino"),
            d.get("entrada_saida"), d.get("nro_unico"), d.get("numero_nota"),
            t.origem.cfop_int, d.get("cfop_descricao"), d.get("cst"), d.get("especie"),
            t.origem.produto_codigo, d.get("produto_descricao"),
            d.get("valor_contabil"), d.get("base_icms"), d.get("aliquota_icms"),
            d.get("valor_icms"),
            t.carga.carga_bruta, t.carga.carga, t.carga.situacao.value,
            t.classificacao.categoria, t.carga.regra, t.classificacao.regra,
            "; ".join(t.alertas), "; ".join(t.pendencias),
            # O que foi informado volta como está: realimentar o arquivo não
            # pode apagar o ajuste de quem o escreveu.
            d.get("ajuste_linha"), d.get("ajuste_valor"), d.get("ajuste_motivo"),
            d.get("ajuste_responsavel"), d.get("ajuste_aprovador"),
            *(d.get(campo) for campo in demais),
        ])
    primeira_moeda = len(COLUNAS_DE_PROCEDENCIA) + 13      # valor_contabil
    for linha in aba.iter_rows(min_row=2, min_col=primeira_moeda,
                               max_col=primeira_moeda + 3):
        for celula in linha:
            celula.number_format = MOEDA
    for linha in aba.iter_rows(min_row=2, min_col=PRIMEIRA_DE_AJUSTE + 1,
                               max_col=PRIMEIRA_DE_AJUSTE + 1):
        for celula in linha:
            celula.number_format = MOEDA
    aba.auto_filter.ref = aba.dimensions


def _aba_pendencias(wb, base: BaseTratada, apuracao: Apuracao) -> None:
    """Tudo que bloqueia o encerramento, inclusive o que não vem de uma linha.

    Atividade indefinida é pendência da apuração, não da base: ela não tem
    linha de origem, e por isso já ficou de fora daqui uma vez. Quem trabalha
    pela planilha não pode deixar de ver um bloqueio que a tela mostra.
    """
    aba = wb.create_sheet("PENDÊNCIAS")
    colunas = [
        ("linha_origem", 12), ("estabelecimento", 30), ("nro_unico", 12),
        ("cfop", 8), ("produto", 12), ("produto_descricao", 38),
        ("valor_icms", 15), ("pendencia", 90),
    ]
    _escreve_cabecalho(aba, colunas)
    for t in base.com_pendencia:
        d = t.origem.dados
        aba.append([
            t.origem.linha_origem, d.get("estabelecimento"), d.get("nro_unico"),
            t.origem.cfop_int, t.origem.produto_codigo, d.get("produto_descricao"),
            d.get("valor_icms"), " | ".join(t.pendencias),
        ])
    for filial in apuracao.sem_regra_de_atividade:
        somas = filial.atividades_sem_regra
        aba.append([
            "", filial.estabelecimento, "", "", "", "", somas.credito_bruto,
            f"ATIVIDADE INDEFINIDA: {somas.linhas} linha(s) não casaram com "
            "nenhuma atividade — cadastre o CFOP em regimes.yaml, bloco "
            "`atividades`",
        ])
    for motivo in apuracao.bloqueios_de_ajuste:
        aba.append(["", "", "", "", "", "", None, f"AJUSTE INCOMPLETO: {motivo}"])
    for linha in aba.iter_rows(min_row=2, min_col=7, max_col=7):
        for celula in linha:
            celula.number_format = MOEDA
    aba.auto_filter.ref = aba.dimensions


def _aba_resumo(wb, base: BaseTratada) -> None:
    aba = wb.create_sheet("RESUMO")
    resumo = base.resumo()
    aba.column_dimensions["A"].width = 34
    aba.column_dimensions["B"].width = 30
    aba.column_dimensions["C"].width = 18
    aba.column_dimensions["D"].width = 18

    def secao(titulo: str) -> None:
        aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
        aba.append([titulo])
        celula = aba.cell(row=aba.max_row, column=1)
        celula.font, celula.fill = TITULO, FUNDO

    aba.append(["APURAÇÃO DE ICMS — BASE TRATADA"])
    aba.cell(row=1, column=1).font = Font(bold=True, size=14)

    secao("PROCEDÊNCIA")
    aba.append(["Versão do Apurabot", __version__])
    for rotulo, chave in [
        ("Competência", "competencia"), ("Período do movimento", "periodo"),
        ("Arquivo de origem", "arquivo"),
        ("SHA-256 do arquivo", "sha256"), ("Gerado em", "gerado_em"),
    ]:
        aba.append([rotulo, str(resumo[chave])])

    secao("VOLUME")
    for rotulo, chave in [
        ("Linhas no Livro Fiscal", "linhas_no_livro"),
        ("Linhas relevantes para ICMS", "linhas_relevantes"),
        ("Pendências (bloqueiam o encerramento)", "pendencias"),
        ("Alertas (não bloqueiam)", "alertas"),
    ]:
        aba.append([rotulo, resumo[chave]])
    aba.append([
        "Encerramento da competência",
        "LIBERADO" if base.pode_encerrar else "BLOQUEADO por pendência",
    ])
    aba.cell(row=aba.max_row, column=2).font = Font(
        bold=True, color="1E7B34" if base.pode_encerrar else "B00020"
    )

    secao("SITUAÇÃO DA EQUALIZAÇÃO")
    for nome, n in sorted(resumo["situacoes"].items(), key=lambda kv: -kv[1]):
        aba.append([nome, n])

    secao("CATEGORIAS DA OPERAÇÃO")
    aba.append(["categoria", "linhas", "ICMS (R$)"])
    for nome, dados in sorted(
        resumo["categorias"].items(), key=lambda kv: -kv[1]["valor_icms"]
    ):
        aba.append([nome, dados["linhas"], dados["valor_icms"]])
        aba.cell(row=aba.max_row, column=3).number_format = MOEDA


def _aba_por_carga(wb, base: BaseTratada) -> None:
    aba = wb.create_sheet("POR ESTABELECIMENTO E CARGA")
    colunas = [
        ("estabelecimento", 30), ("entrada_saida", 14), ("carga_efetiva", 14),
        ("linhas", 10), ("valor_contabil", 18), ("base_icms", 18), ("valor_icms", 18),
    ]
    _escreve_cabecalho(aba, colunas)
    somas = base.por_estabelecimento_carga()
    for (estab, es, carga), v in sorted(
        somas.items(), key=lambda kv: (str(kv[0][0]), str(kv[0][1]), str(kv[0][2]))
    ):
        aba.append([
            estab, es, "(vazio)" if carga is None else carga, v["linhas"],
            v["valor_contabil"], v["base_icms"], v["valor_icms"],
        ])
    for linha in aba.iter_rows(min_row=2, min_col=5, max_col=7):
        for celula in linha:
            celula.number_format = MOEDA
    aba.auto_filter.ref = aba.dimensions


def _aba_apuracao(wb, apuracao: Apuracao) -> None:
    aba = wb.create_sheet("APURAÇÃO POR FILIAL")
    colunas = [
        ("estabelecimento", 32), ("uf", 6), ("regime", 28), ("linhas", 9),
        ("credito_bruto", 16), ("estorno", 16), ("credito_indevido", 17),
        ("credito_mantido", 17), ("debito", 16), ("credito_presumido", 18),
        ("difal", 14), ("saldo credor anterior", 21), ("saldo (credor +)", 18),
        ("a recolher", 15), ("confere", 10),
    ]
    _escreve_cabecalho(aba, colunas)
    primeira = aba.max_row + 1
    for f in sorted(apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)):
        aba.append([
            f.estabelecimento, f.uf, f.regime, f.linhas, f.credito_bruto,
            f.estorno, f.credito_indevido, f.credito_mantido, f.debito,
            f.credito_presumido, f.difal, f.saldo_credor_anterior, f.saldo,
            f.a_recolher, "OK" if f.confere else "DIVERGE",
        ])
    ultima = aba.max_row
    total = apuracao.total
    aba.append([
        "TOTAL", "", "", total.linhas, total.credito_bruto, total.estorno,
        total.credito_indevido, total.credito_mantido, total.debito,
        total.credito_presumido, total.difal, total.saldo_credor_anterior,
        total.saldo, apuracao.a_recolher, "OK" if total.confere else "DIVERGE",
    ])
    # O TOTAL soma as filiais na própria planilha, em vez de repetir o número
    # que o motor calculou — inclusive "a recolher", que é a soma do que cada
    # uma paga e não o saldo do grupo com o sinal trocado: filial credora não
    # paga a conta de outra devedora fora da centralização.
    linha = aba.max_row
    if ultima >= primeira:
        for coluna in range(4, 15):
            letra = get_column_letter(coluna)
            aba.cell(row=linha, column=coluna).value = (
                f"=SUM({letra}{primeira}:{letra}{ultima})"
            )
    for celula in aba[linha]:
        celula.font = Font(bold=True)
    for linha in aba.iter_rows(min_row=2, min_col=5, max_col=14):
        for celula in linha:
            celula.number_format = MOEDA

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append([
        "Saldo na convenção de caixa: positivo é credor — crédito a transportar "
        "—, negativo é devedor. \"A recolher\" é o que sai do caixa, e o TOTAL "
        "dela é a soma das filiais, não o saldo do grupo com o sinal trocado. "
        "O saldo é o FINAL, o mesmo que o Registro de cada estabelecimento "
        "fecha: já abre com o crédito do mês anterior e já traz o efeito da "
        "centralização. O saldo antes de centralizar está no bloco de "
        "Centralização, mais abaixo."
    ])

    _bloco_saldo_credor(aba, apuracao)
    _bloco_ajustes(aba, apuracao)

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Memória do benefício fiscal"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    for f in sorted(apuracao.filiais.values(), key=lambda f: f.estabelecimento):
        if not f.beneficio:
            continue
        aba.append([f.estabelecimento])
        aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
        for passo in f.beneficio.memoria:
            aba.append(["", passo])

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Centralização e transferência de saldo"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    for c in apuracao.centralizacao:
        for passo in c.memoria:
            aba.append(["", passo])
    aba.append(["", "As transferências a emitir estão na aba TRANSFERÊNCIAS."])

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Contribuição ao Pró-Desenvolve / FADEFE — GUIA AVULSA"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    aba.append(["", "Informativo: não entra na conta gráfica da apuração."])
    aba.append(["estabelecimento", "benefício fruído", "%", "a recolher",
                "% adicional", "adicional a recolher"])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    primeira = aba.max_row + 1
    for f in sorted(apuracao.filiais.values(), key=lambda f: f.estabelecimento):
        if not f.beneficio or not f.beneficio.percentual_fadefe:
            continue
        b = f.beneficio
        aba.append([
            f.estabelecimento, b.credito_presumido, b.percentual_fadefe / 100,
            b.fadefe, b.percentual_fadefe_adicional / 100, b.fadefe_adicional,
        ])
    for linha in aba.iter_rows(min_row=primeira, max_row=aba.max_row,
                               min_col=2, max_col=6):
        for celula in linha:
            celula.number_format = PERCENTUAL if celula.column in (3, 5) else MOEDA

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Segregação por atividade (exigida pela GIA de MS)"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    aba.append(["estabelecimento", "atividade", "linhas", "credito_bruto",
                "estorno", "credito_mantido", "debito", "debito_intra",
                "debito_inter", "saldo"])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    primeira_atividade = aba.max_row + 1
    for f in sorted(apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)):
        if not f.segrega_por_atividade:
            continue
        conhecidas = [a for a in ATIVIDADES_EM_ORDEM if a in f.por_atividade]
        restantes = [a for a in f.por_atividade if a not in conhecidas]
        for nome in conhecidas + sorted(restantes):
            t_ = f.por_atividade[nome]
            aba.append([
                f.estabelecimento, nome, t_.linhas, t_.credito_bruto, t_.estorno,
                t_.credito_mantido, t_.debito, t_.debito_de(INTRAESTADUAL),
                t_.debito_de(INTERESTADUAL), t_.saldo,
            ])
    for linha in aba.iter_rows(min_row=primeira_atividade, max_row=aba.max_row,
                               min_col=4, max_col=10):
        for celula in linha:
            celula.number_format = MOEDA

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Detalhe por carga efetiva"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    aba.append(["estabelecimento", "carga", "credito_bruto", "estorno",
                "credito_indevido", "credito_mantido"])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    for f in sorted(apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)):
        for carga, v in sorted(f.por_carga.items(), key=lambda kv: str(kv[0])):
            aba.append([
                f.estabelecimento, carga, v["credito_bruto"], v["estorno"],
                v["credito_indevido"], v["credito_mantido"],
            ])
            for coluna in (3, 4, 5, 6):
                aba.cell(row=aba.max_row, column=coluna).number_format = MOEDA


def _bloco_ajustes(aba, apuracao: Apuracao) -> None:
    """A memória dos ajustes: o que entrou, de onde veio e quem aprovou.

    Junta as duas origens numa lista só — a aba AJUSTES é o formulário, esta é
    a prestação de contas. E traz o que ficou marcado sem ser lançado, que não
    muda número nenhum mas não pode sumir.
    """
    a = apuracao.ajustes
    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Ajustes declarados"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    if not a.lancamentos:
        aba.append(["", "Nenhum. As linhas 002, 003, 006 e 007 do registro saem "
                        "marcadas até alguém assinar a conferência."])
    else:
        aba.append(["estabelecimento", "atividade", "linha", "valor", "motivo",
                    "responsável", "aprovador", "onde foi informado"])
        for celula in aba[aba.max_row]:
            celula.font, celula.fill = TITULO, FUNDO
        primeira = aba.max_row + 1
        for x in a.lancamentos:
            aba.append([x.estabelecimento, x.atividade, f"{x.linha:03d}", x.valor,
                        x.motivo, x.responsavel, x.aprovador, x.onde])
        for linha in aba.iter_rows(min_row=primeira, max_row=aba.max_row,
                                   min_col=4, max_col=4):
            for celula in linha:
                celula.number_format = MOEDA

    if a.anotacoes:
        aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
        aba.append([f"Marcado, não lançado — {len(a.anotacoes)} linha(s)"])
        aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
        aba.append(["", "Não entra na apuração; fica aqui para não se perder."])
        aba.append(["estabelecimento", "", "", "valor", "motivo", "responsável",
                    "", "onde foi informado"])
        for celula in aba[aba.max_row]:
            celula.font, celula.fill = TITULO, FUNDO
        primeira = aba.max_row + 1
        for x in a.anotacoes:
            aba.append([x.estabelecimento, "", "", x.valor, x.motivo,
                        x.responsavel, "", x.onde])
        for linha in aba.iter_rows(min_row=primeira, max_row=aba.max_row,
                                   min_col=4, max_col=4):
            for celula in linha:
                celula.number_format = MOEDA
        aba.append(["", "Total marcado e não lançado", "", a.marcado_nao_lancado])
        aba.cell(row=aba.max_row, column=4).number_format = MOEDA
        aba.cell(row=aba.max_row, column=2).font = Font(bold=True)


def _bloco_saldo_credor(aba, apuracao: Apuracao) -> None:
    """A conta gráfica atravessa a virada do mês — e a planilha mostra por onde.

    A última coluna é o que a competência seguinte tem que receber como
    abertura. Sai daqui pronta para o cadastro, para ninguém ter que subtrair
    duas linhas do registro à mão.
    """
    anterior, seguinte = apuracao.competencia_anterior, apuracao.competencia_seguinte
    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["Saldo credor — linhas 009 e 014 do Registro de Apuração"])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    if not apuracao.saldos_declarados:
        aba.append([
            "",
            f"A abertura de {apuracao.competencia} não está declarada em "
            "parametros/saldos.yaml — a apuração rodou com todos os "
            "estabelecimentos abrindo o mês zerados.",
        ])
    aba.append([
        "estabelecimento", f"veio de {anterior}", "apurado no mês",
        f"vai para {seguinte}",
    ])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    primeira = aba.max_row + 1
    for f in sorted(apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)):
        if not (f.saldo_credor_anterior or f.credor):
            continue
        aba.append([f.estabelecimento, f.saldo_credor_anterior, f.saldo_do_periodo,
                    f.credor])
    for linha in aba.iter_rows(min_row=primeira, max_row=aba.max_row,
                               min_col=2, max_col=4):
        for celula in linha:
            celula.number_format = MOEDA
    aba.append([
        "",
        f"A coluna \"vai para {seguinte}\" é a abertura da competência seguinte: "
        "cadastre-a em parametros/saldos.yaml.",
    ])


def _aba_ajustes(wb, apuracao: Apuracao) -> None:
    """O formulário dos ajustes que não têm documento, e a conferência.

    Duas coisas moram aqui. As **parcelas sem documento** são os lançamentos do
    Registro que não pertencem a nota nenhuma — o ajuste que tem dono vai na
    linha dele, na BASE TRATADA, e nem passa por aqui.

    A **conferência** é o que faz a marca `AGUARDA AJUSTE` sumir. Célula vazia
    diz ao mesmo tempo "não tem ajuste" e "ninguém olhou ainda", e a ferramenta
    não tem como escolher uma das duas: por isso alguém assina, e assinar sem
    nenhum ajuste também é resposta.
    """
    from .ajustes import ABA, TITULO_CONFERENCIA, TITULO_PARCELAS

    aba = wb.create_sheet(ABA)
    for coluna, largura in zip("ABCDEFG", (32, 22, 10, 16, 52, 22, 22)):
        aba.column_dimensions[coluna].width = largura

    aba.append([f"AJUSTES DA APURAÇÃO — competência {apuracao.competencia}"])
    aba.cell(row=1, column=1).font = Font(bold=True, size=14)
    for texto in (
        "Preencha, salve e arraste este mesmo arquivo de volta no Apurabot.",
        "O valor é sempre positivo: quem dá o sentido é a linha do Registro — "
        "002 e 003 aumentam o que se deve, 006 e 007 diminuem.",
        "Use ANOTAR na coluna `linha` para marcar sem lançar (ICMS em "
        "discussão, por exemplo): não muda a apuração e sai no relatório.",
        "Ajuste que pertence a uma nota vai na linha dela, na aba BASE "
        "TRATADA — aqui só o que não tem documento.",
    ):
        aba.append(["", texto])

    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append([TITULO_PARCELAS])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    aba.append(["estabelecimento", "atividade", "linha", "valor", "motivo",
                "responsável", "aprovador"])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    primeira = aba.max_row + 1
    for ajuste in apuracao.ajustes.lancamentos + apuracao.ajustes.anotacoes:
        if ajuste.onde.startswith("BASE TRATADA"):
            continue                    # esse tem dono; mora na linha dele
        aba.append([
            ajuste.estabelecimento, ajuste.atividade,
            "ANOTAR" if ajuste.anotacao else f"{ajuste.linha:03d}",
            ajuste.valor, ajuste.motivo, ajuste.responsavel, ajuste.aprovador,
        ])
    for _ in range(12):                 # espaço para escrever
        aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    for linha in aba.iter_rows(min_row=primeira, max_row=aba.max_row,
                               min_col=4, max_col=4):
        for celula in linha:
            celula.number_format = MOEDA

    aba.append([TITULO_CONFERENCIA])
    aba.cell(row=aba.max_row, column=1).font = Font(bold=True)
    aba.append(["estabelecimento", "conferido por", "conferido em", "observação"])
    for celula in aba[aba.max_row]:
        celula.font, celula.fill = TITULO, FUNDO
    for filial in sorted(
        apuracao.filiais.values(), key=lambda f: (f.uf, f.estabelecimento)
    ):
        assinada = apuracao.ajustes.conferencia.get(filial.estabelecimento)
        aba.append([
            filial.estabelecimento,
            assinada.por if assinada else None,
            assinada.em if assinada else None,
            assinada.observacao if assinada else None,
        ])
    aba.append([None])   # linha em branco: `append([])` não avança no openpyxl
    aba.append(["", "Estabelecimento com `conferido por` preenchido para de "
                    "mostrar AGUARDA AJUSTE — inclusive sem nenhum ajuste."])


#: Ordem de leitura das abas — da conclusão para o detalhe.
ORDEM_DAS_ABAS = [
    "RESUMO", "REGISTRO", "AJUSTES", "APURAÇÃO EFETIVA", "APURAÇÃO POR FILIAL",
    "TRANSFERÊNCIAS", "PENDÊNCIAS", "POR ESTABELECIMENTO E CARGA", "BASE TRATADA",
]


def _ordenar_abas(wb) -> None:
    posicao = {nome: i for i, nome in enumerate(ORDEM_DAS_ABAS)}
    wb._sheets.sort(key=lambda aba: posicao.get(aba.title, len(posicao)))


def escrever(
    base: BaseTratada,
    destino: Path | str,
    apuracao: Apuracao | None = None,
    ajustes=None,
) -> Path:
    """Grava a base tratada, a apuração e as conferências em .xlsx."""
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    apuracao = apuracao if apuracao is not None else apurar(base)

    wb = Workbook()
    wb.remove(wb.active)
    _aba_base(wb, base)
    _aba_pendencias(wb, base, apuracao)
    _aba_por_carga(wb, base)
    _aba_apuracao(wb, apuracao)
    aba_apuracao_efetiva(wb, apuracao, base.parametros)
    aba_registro(wb, apuracao, base.parametros, ajustes)
    _aba_ajustes(wb, apuracao)
    aba_transferencias(wb, apuracao)
    _aba_resumo(wb, base)
    _ordenar_abas(wb)
    wb.save(destino)
    return destino
