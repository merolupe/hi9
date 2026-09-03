"""Os ajustes: o que a apuração decide e o Livro Fiscal não sabe.

Duas portas para as mesmas quatro linhas do Registro. O ajuste que pertence a
um documento vai na linha dele, na BASE TRATADA; o que não pertence a nenhum vai
na aba AJUSTES. O alvo é o Registro de Apuração de Rio Brilhante em 07/2026, que
declara R$ 335.101,41 na linha 003 — R$ 331.236,11 que a regra calcula sobre o
Livro mais R$ 3.865,30 de ajuste.
"""
from __future__ import annotations

import openpyxl
import pytest

from apurabot import ajustes as aj
from apurabot.apuracao import apurar, ler_ajustes
from apurabot.base_tratada import tratar
from apurabot.nucleo import registro as reg
from apurabot.saida import escrever

CENTAVO = 0.005
RB = "HINOVE (RIO BRILHANTE)"
REGISTRO = "HINOVE (REGISTRO)"

AJUSTE_ESTORNO = 3_865.30
AJUSTE_ART_68 = 46_138.68
AJUSTE_ESTORNO_DEBITO = 33_039.71


# -- o que se aceita e o que se recusa --------------------------------------

def test_o_valor_e_sempre_positivo_e_quem_da_o_sentido_e_a_linha():
    """Negativo na 003 seria o mesmo que positivo na 006 — duas formas de
    escrever a mesma coisa tornam a conferência mais difícil sem ganhar nada."""
    with pytest.raises(aj.AjusteInvalido, match="sempre positivo"):
        aj.montar("X", linha="003", valor=-5, motivo="m", responsavel="r",
                  aprovador="a")


def test_linha_fora_das_quatro_nao_existe():
    with pytest.raises(aj.AjusteInvalido, match="não existe"):
        aj.montar("X", linha="009", valor=1, motivo="m", responsavel="r",
                  aprovador="a")


@pytest.mark.parametrize(
    "faltando, esperado",
    [
        ({"motivo": None}, "ajuste_motivo"),
        ({"responsavel": None}, "ajuste_responsavel"),
        ({"aprovador": None}, "ajuste_aprovador"),
        ({"valor": None}, "ajuste_valor"),
        ({"linha": None}, "ajuste_linha"),
    ],
)
def test_ajuste_pela_metade_e_recusado(faltando, esperado):
    """Completar por conta própria é o que a regra 4 do repositório proíbe."""
    campos = dict(linha="003", valor=100.0, motivo="m", responsavel="r",
                  aprovador="a")
    campos.update(faltando)
    with pytest.raises(aj.AjusteInvalido, match=esperado):
        aj.montar("X", **campos)


def test_linha_em_branco_nao_e_ajuste():
    """São milhares de linhas sem ajuste nenhum: elas não podem virar erro."""
    assert aj.montar("X", None, None, None, None, None) is None


def test_o_valor_chega_como_o_excel_e_a_pessoa_escrevem():
    for escrito in ("3.865,30", "3865.30", 3865.3, " R$ 3.865,30 "):
        ajuste = aj.montar("X", "003", escrito, "m", "r", "a")
        assert ajuste.valor == pytest.approx(AJUSTE_ESTORNO, abs=CENTAVO)


def test_anotar_marca_sem_lancar():
    """"Este ICMS é indevido e será tratado por anuência" não move a apuração."""
    ajuste = aj.montar("X", aj.ANOTAR, 12_480.0, "indevido, por anuência", "r", "")
    assert ajuste.anotacao
    assert ajuste.linha is None
    assert ajuste.valor == pytest.approx(12_480.0)


def test_anotacao_dispensa_aprovador_mas_nao_o_motivo():
    """Ela não muda número; o que ela precisa dizer é por que existe."""
    assert aj.montar("X", aj.ANOTAR, None, "em discussão", "r", "") is not None
    with pytest.raises(aj.AjusteInvalido, match="ajuste_motivo"):
        aj.montar("X", aj.ANOTAR, 1.0, None, "r", "")


def test_a_atividade_aceita_o_nome_que_o_fiscal_usa():
    assert aj.montar("X", "003", 1, "m", "r", "a", atividade="Produção").atividade == (
        "industrial"
    )
    with pytest.raises(aj.AjusteInvalido, match="não existe"):
        aj.montar("X", "003", 1, "m", "r", "a", atividade="agrícola")


# -- a volta pela planilha --------------------------------------------------

@pytest.fixture(scope="module")
def devolvido(base_julho, parametros, tmp_path_factory):
    """A saída da ferramenta, preenchida como o time fiscal preencheria."""
    destino = tmp_path_factory.mktemp("ajustes") / "Apuracao_2026-07.xlsx"
    escrever(base_julho, destino, apurar(base_julho, parametros))

    wb = openpyxl.load_workbook(destino)
    aba = wb[aj.ABA]
    titulos = {str(l[0].value or "").upper(): l[0].row for l in aba.iter_rows(max_col=1)}

    linha = titulos[aj.TITULO_PARCELAS] + 2
    for i, (codigo, valor, motivo) in enumerate([
        ("003", AJUSTE_ESTORNO, "estorno para ajuste de apuração"),
        ("006", AJUSTE_ART_68, "crédito do art. 68 do RICMS/MS"),
        ("007", AJUSTE_ESTORNO_DEBITO, "estorno de débitos"),
    ]):
        for coluna, v in enumerate(
            [RB, "Produção", codigo, valor, motivo, "Fulano", "Ciclano"], start=1
        ):
            aba.cell(row=linha + i, column=coluna, value=v)

    conferencia = titulos[aj.TITULO_CONFERENCIA] + 2
    for i in range(8):
        if aba.cell(row=conferencia + i, column=1).value:
            aba.cell(row=conferencia + i, column=2, value="Fulano")
            aba.cell(row=conferencia + i, column=3, value="05/09/2026")

    # E um ajuste com dono, na linha da nota, mais uma anotação.
    base = wb["BASE TRATADA"]
    cabecalho = {str(c.value): c.column for c in base[1]}
    alvo = next(
        linha[0].row
        for linha in base.iter_rows(min_row=2)
        if "REGISTRO" in str(linha[cabecalho["estabelecimento"] - 1].value or "")
    )
    for nome, valor in [
        ("ajuste_linha", "003"), ("ajuste_valor", 500.0),
        ("ajuste_motivo", "erro de lançamento"),
        ("ajuste_responsavel", "Fulano"), ("ajuste_aprovador", "Ciclano"),
    ]:
        base.cell(row=alvo, column=cabecalho[nome], value=valor)
    for nome, valor in [
        ("ajuste_linha", aj.ANOTAR), ("ajuste_valor", 12_480.0),
        ("ajuste_motivo", "indevido, tratado por anuência"),
        ("ajuste_responsavel", "Fulano"),
    ]:
        base.cell(row=alvo + 1, column=cabecalho[nome], value=valor)

    devolvido = destino.parent / "devolvido.xlsx"
    wb.save(devolvido)
    return devolvido


@pytest.fixture(scope="module")
def reapurado(devolvido, parametros):
    base = tratar(devolvido, parametros=parametros)
    return base, apurar(base, parametros, ajustes=ler_ajustes(devolvido))


def test_a_ferramenta_le_a_propria_saida(reapurado, base_julho):
    """O arquivo devolvido é autossuficiente: leva o Livro inteiro dentro."""
    base, _ = reapurado
    assert len(base.linhas) == len(base_julho.linhas)
    assert base.competencia == "2026-07"
    assert base.periodo == base_julho.periodo


def test_o_registro_fecha_no_valor_que_o_erp_declara(reapurado, parametros):
    """A prova do caminho inteiro: preencher, devolver, e o livro fechar."""
    _, apuracao = reapurado
    rb = next(
        r for r in reg.montar(apuracao, parametros) if r.estabelecimento == RB
    )
    esperado = {1: 505_991.41, 2: 99_412.10, 3: 335_101.41, 4: 940_504.92,
                5: 469_903.05, 8: 549_081.44, 10: 549_081.44, 11: 391_423.48}
    for codigo, valor in esperado.items():
        assert rb.linha(codigo).valor == pytest.approx(valor, abs=CENTAVO), codigo


def test_as_duas_origens_somam_na_mesma_linha(reapurado):
    _, apuracao = reapurado
    de_onde = {a.onde.split(",")[0] for a in apuracao.ajustes.lancamentos}
    assert de_onde == {"aba AJUSTES", "BASE TRATADA"}
    assert len(apuracao.ajustes.lancamentos) == 4


def test_o_ajuste_da_linha_nao_precisa_dizer_de_quem_e(reapurado):
    """Estabelecimento e atividade saem da linha — é a vantagem de morar nela."""
    _, apuracao = reapurado
    da_linha = next(
        a for a in apuracao.ajustes.lancamentos if a.onde.startswith("BASE TRATADA")
    )
    assert da_linha.estabelecimento == REGISTRO
    assert da_linha.valor == pytest.approx(500.0)


def test_o_ajuste_muda_o_saldo_da_apuracao_e_nao_so_o_registro(reapurado):
    """Se mudasse só o registro, a tela e o livro contariam histórias diferentes."""
    _, apuracao = reapurado
    registro = apuracao.filiais[REGISTRO]
    # -299.450,70 do mês (já com o DIFAL de SP) menos os 500,00 do ajuste.
    assert registro.saldo_individual == pytest.approx(-299_950.70, abs=CENTAVO)
    # O saldo final é zero: Registro é centralizado em SP e transfere tudo
    # para Guará. O ajuste mudou o que ele transfere, não o que ele recolhe.
    assert registro.saldo == 0.0
    assert registro.a_recolher == 0.0
    assert registro.efeito_da_centralizacao == pytest.approx(
        299_950.70, abs=CENTAVO
    )


def test_anotar_nao_entra_na_conta_mas_aparece(reapurado):
    _, apuracao = reapurado
    assert apuracao.ajustes.marcado_nao_lancado == pytest.approx(12_480.0)
    assert len(apuracao.ajustes.anotacoes) == 1
    # e o estabelecimento anotado não teve o saldo mexido pela anotação
    assert 3 not in [a.linha for a in apuracao.ajustes.anotacoes]


def test_a_conferencia_assinada_tira_a_marca_do_registro(reapurado, parametros):
    _, apuracao = reapurado
    for registro in reg.montar(apuracao, parametros):
        assert not registro.aguarda_ajustes, registro.estabelecimento


def test_sem_conferencia_a_marca_continua(base_julho, parametros):
    apuracao = apurar(base_julho, parametros)
    rb = next(r for r in reg.montar(apuracao, parametros) if r.estabelecimento == RB)
    assert rb.aguarda_ajustes
    assert {i.codigo for i in rb.resumo if i.aguarda_ajuste} == {3, 6, 7}


def test_o_ajuste_escrito_volta_na_planilha(reapurado, tmp_path):
    """Realimentar não pode apagar o ajuste de quem o escreveu."""
    base, apuracao = reapurado
    destino = tmp_path / "de_novo.xlsx"
    escrever(base, destino, apuracao)
    aba = openpyxl.load_workbook(destino)["BASE TRATADA"]
    cabecalho = {str(c.value): c.column for c in aba[1]}
    motivos = {
        aba.cell(row=linha[0].row, column=cabecalho["ajuste_motivo"]).value
        for linha in aba.iter_rows(min_row=2)
    }
    assert "erro de lançamento" in motivos
    assert "indevido, tratado por anuência" in motivos


# -- o que fica bloqueado ---------------------------------------------------

def test_ajuste_incompleto_na_linha_vira_pendencia(base_julho, parametros, tmp_path):
    """Ignorar seria perder uma decisão que alguém tomou."""
    destino = tmp_path / "meio.xlsx"
    escrever(base_julho, destino, apurar(base_julho, parametros))
    wb = openpyxl.load_workbook(destino)
    base = wb["BASE TRATADA"]
    cabecalho = {str(c.value): c.column for c in base[1]}
    base.cell(row=2, column=cabecalho["ajuste_valor"], value=1000.0)
    quebrado = tmp_path / "quebrado.xlsx"
    wb.save(quebrado)

    tratada = tratar(quebrado, parametros=parametros)
    assert not tratada.pode_encerrar
    assert any("AJUSTE INCOMPLETO" in p for t in tratada.com_pendencia
               for p in t.pendencias)


def test_parcela_sem_atividade_onde_a_uf_segrega_bloqueia(base_julho, parametros):
    """Em MS a atividade dimensiona o benefício: lançar sem ela move o incentivo."""
    declarados = aj.Declarados(parcelas=[
        aj.Ajuste(estabelecimento=RB, linha=3, valor=1_000.0, motivo="m",
                  responsavel="r", aprovador="a", onde="aba AJUSTES, linha 9")
    ])
    from apurabot.apuracao import de_declarados

    apuracao = apurar(base_julho, parametros, ajustes=de_declarados(declarados))
    assert apuracao.ajustes_sem_atividade
    assert any("segrega por atividade" in m for m in apuracao.bloqueios_de_ajuste)


def test_onde_a_uf_nao_segrega_a_atividade_nao_faz_falta(base_julho, parametros):
    from apurabot.apuracao import de_declarados

    declarados = aj.Declarados(parcelas=[
        aj.Ajuste(estabelecimento=REGISTRO, linha=3, valor=1_000.0, motivo="m",
                  responsavel="r", aprovador="a", onde="aba AJUSTES, linha 9")
    ])
    apuracao = apurar(base_julho, parametros, ajustes=de_declarados(declarados))
    assert not apuracao.bloqueios_de_ajuste


def test_apurar_duas_vezes_nao_dobra_o_ajuste(base_julho, parametros):
    """A apuração acrescenta os ajustes de linha — no objeto dela, não no seu."""
    from apurabot.apuracao import AjustesDaApuracao

    meus = AjustesDaApuracao()
    apurar(base_julho, parametros, ajustes=meus)
    apurar(base_julho, parametros, ajustes=meus)
    assert meus.lancamentos == []
