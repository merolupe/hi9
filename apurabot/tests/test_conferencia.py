"""As abas de conferência: o que elas prometem ao time fiscal."""
from __future__ import annotations

import re

import openpyxl
import pytest

from apurabot.apuracao import apurar
from apurabot.conferencia import (
    ROTULO_DA_ATIVIDADE,
    ROTULO_DA_CATEGORIA,
    _rotulo_atividade,
    rotulo_da_categoria,
)
from apurabot.nucleo import atividade as ativ
from apurabot.saida import ORDEM_DAS_ABAS, escrever

CENTAVO = 0.005


@pytest.fixture(scope="module")
def apuracao(base_julho):
    return apurar(base_julho)


@pytest.fixture(scope="module")
def planilha(base_julho, apuracao, tmp_path_factory):
    destino = tmp_path_factory.mktemp("saida") / "conferencia.xlsx"
    escrever(base_julho, destino, apuracao)
    return openpyxl.load_workbook(destino, data_only=True)


# -- o CHECK é a promessa da aba -------------------------------------------

def test_toda_linha_apurada_fecha_o_check(apuracao):
    """a estornar + a apropriar = ICMS creditado, linha a linha.

    É o que a coluna CHECK mostra. Se falhar aqui, a aba mostra vermelho — e
    o vermelho é erro de motor, não de escrituração.
    """
    quebradas = [
        a for f in apuracao.filiais.values() for a in f.apuradas if not a.confere
    ]
    assert not quebradas, f"{len(quebradas)} linha(s) não fecham o CHECK"


def test_a_soma_das_linhas_apuradas_e_o_total_da_filial(apuracao):
    for filial in apuracao.filiais.values():
        assert sum(a.resultado.credito_bruto for a in filial.apuradas) == pytest.approx(
            filial.credito_bruto, abs=CENTAVO
        )
        assert sum(a.credito_a_apropriar for a in filial.apuradas) == pytest.approx(
            filial.credito_mantido, abs=CENTAVO
        )
        assert sum(a.resultado.debito for a in filial.apuradas) == pytest.approx(
            filial.debito, abs=CENTAVO
        )


def test_a_conferencia_agrupa_sempre_pela_carga_efetiva(planilha):
    """Os dois regimes agrupam pela carga efetiva equalizada.

    Antes MS agrupava pela alíquota, e a mesma aba trazia duas grandezas
    diferentes na mesma coluna. A carga efetiva é o que o documento traz depois
    da equalização, e é por ela que a conferência manual olha nos dois estados.
    """
    aba = planilha["APURAÇÃO EFETIVA"]
    cabecalhos = {
        str(linha[2].value) for linha in aba.iter_rows(max_col=4)
        if str(linha[0].value or "") == "CFOP"
    }
    assert cabecalhos == {"Carga efetiva"}, cabecalhos


def test_a_aba_tem_uma_coluna_de_percentual_e_nenhum_check(planilha):
    """Uma coluna de percentual, sempre preenchida. Sem coluna de CHECK.

    A `% da regra` saía vazia nas linhas de CFOP que misturam cargas, porque
    ali não existe um nominal só. Meia coluna preenchida confunde mais do que
    informa. O CHECK saiu porque a identidade que ele mostrava é garantida por
    `test_toda_linha_apurada_fecha_o_check`, no motor — não por uma célula que
    alguém precise conferir.
    """
    aba = planilha["APURAÇÃO EFETIVA"]
    cabecalhos = [
        linha for linha in aba.iter_rows(max_col=11, values_only=True)
        if str(linha[0] or "") == "CFOP"
    ]
    assert cabecalhos
    for linha in cabecalhos:
        assert "CHECK" not in linha
        percentuais = [c for c in linha if "%" in str(c or "")]
        assert percentuais == ["% do crédito estornado"], linha


def test_a_parcela_de_ms_e_a_formula_da_regra(apuracao):
    """Em MS, estorno ÷ crédito é exatamente 1 − 4/alíquota.

    É a única UF em que o percentual da conferência coincide com um parâmetro
    da regra — por isso o rótulo da coluna é descritivo, e não normativo.
    """
    rb = apuracao.filiais["HINOVE (RIO BRILHANTE)"]
    conferidas = 0
    for a in rb.apuradas:
        if not a.resultado.credito_bruto or not a.credito_a_estornar:
            continue
        aliquota = float(a.tratada.origem.dados.get("aliquota_icms") or 0)
        if aliquota <= 4:
            continue
        assert a.credito_a_estornar / a.resultado.credito_bruto == pytest.approx(
            round(1 - 4 / aliquota, 4), abs=1e-6
        )
        conferidas += 1
    assert conferidas, "nenhuma linha de RB com estorno proporcional"


def test_em_sp_o_percentual_nao_e_parametro_da_regra(apuracao):
    """Em SP o estorno incide sobre o valor contábil, não sobre o ICMS.

    A razão estorno ÷ crédito varia dentro de uma mesma carga, porque a carga
    do documento foi equalizada para a régua nominal. Chamar isso de "parcela
    não tributada" — que é vocabulário de MS — ensinaria a regra errada.
    """
    guara = apuracao.filiais["HINOVE (FILIAL GUARÁ)"]
    razoes = {}
    for a in guara.apuradas:
        if not a.resultado.credito_bruto or not a.credito_a_estornar:
            continue
        carga = a.tratada.carga.carga
        razoes.setdefault(carga, set()).add(
            round(a.credito_a_estornar / a.resultado.credito_bruto, 4)
        )
    assert any(len(v) > 1 for v in razoes.values()), (
        "esperava razões diferentes dentro de uma mesma carga em SP"
    )


# -- rótulos ----------------------------------------------------------------

def test_a_atividade_aparece_com_o_nome_que_o_fiscal_usa(apuracao):
    rb = apuracao.filiais["HINOVE (RIO BRILHANTE)"]
    industriais = [a for a in rb.apuradas if a.atividade == ativ.INDUSTRIAL]
    assert industriais
    assert _rotulo_atividade(industriais[0]) == "Produção"
    assert ROTULO_DA_ATIVIDADE[ativ.COMERCIAL] == "Comercial"


def test_onde_a_uf_nao_segrega_o_rotulo_e_a_categoria_da_equalizacao(apuracao):
    guara = apuracao.filiais["HINOVE (FILIAL GUARÁ)"]
    assert not guara.segrega_por_atividade
    rotulos = {_rotulo_atividade(a) for a in guara.apuradas}
    assert rotulos
    assert not rotulos & set(ROTULO_DA_ATIVIDADE.values())
    # Nome de gente, não nome de campo.
    assert "Matéria-Prima" in rotulos
    assert not any("_" in r for r in rotulos), rotulos


def test_a_categoria_sai_com_nome_legivel(apuracao):
    """`frete_transferencia` na base, "Frete de Transferência" na conferência."""
    assert rotulo_da_categoria("frete_transferencia") == "Frete de Transferência"
    assert rotulo_da_categoria("materia_prima") == "Matéria-Prima"
    assert rotulo_da_categoria("ciap") == "CIAP"
    assert rotulo_da_categoria("") == ""

    # Categoria fora do mapa não sai com cara de código.
    assert rotulo_da_categoria("credito_de_teste") == "Credito de Teste"

    # Toda categoria que a competência produz tem rótulo cadastrado.
    usadas = {
        a.tratada.classificacao.categoria
        for f in apuracao.filiais.values() for a in f.apuradas
        if a.tratada.classificacao.categoria
    }
    assert usadas <= set(ROTULO_DA_CATEGORIA), usadas - set(ROTULO_DA_CATEGORIA)


# -- a planilha -------------------------------------------------------------

def test_a_planilha_traz_as_tres_abas_de_conferencia(planilha):
    assert {"APURAÇÃO EFETIVA", "REGISTRO", "TRANSFERÊNCIAS"} <= set(
        planilha.sheetnames
    )


def test_as_abas_saem_da_conclusao_para_o_detalhe(planilha):
    """Quem abre o arquivo cai no resumo, não em seis mil linhas de base."""
    assert planilha.sheetnames == [
        nome for nome in ORDEM_DAS_ABAS if nome in planilha.sheetnames
    ]
    assert planilha.sheetnames[0] == "RESUMO"


def test_a_aba_de_transferencias_diz_o_que_emitir(planilha):
    texto = "\n".join(
        str(c.value)
        for linha in planilha["TRANSFERÊNCIAS"].iter_rows()
        for c in linha
        if c.value
    )
    assert "HINOVE (REGISTRO)" in texto and "HINOVE (FILIAL GUARÁ)" in texto
    assert "HINOVE (CORUMBÁ- MS)" in texto and "HINOVE (RIO BRILHANTE)" in texto
    assert "NF-e" in texto and "Registro de Apuração" in texto


def test_a_aba_de_pendencias_nao_cobra_mais_a_nota_de_transferencia(planilha):
    texto = "\n".join(
        str(c.value)
        for linha in planilha["PENDÊNCIAS"].iter_rows()
        for c in linha
        if c.value
    )
    assert "sem NF-e escriturada" not in texto


def test_o_registro_tem_um_bloco_por_filial_e_o_totalizador(planilha, apuracao):
    primeira = [linha[0].value for linha in planilha["REGISTRO"].iter_rows(max_col=1)]
    for nome in apuracao.filiais:
        assert nome in primeira
    assert any(
        str(v or "").startswith("TOTALIZADOR") for v in primeira
    ), "falta o totalizador que o PDF por filial não tem"


# -- o que a rodada de um pré-livro exigiu ----------------------------------

def test_a_planilha_mostra_a_pendencia_de_atividade(base_julho, parametros, tmp_path):
    """Atividade indefinida não tem linha de origem — e por isso ficou de fora.

    Quem trabalha pela planilha não pode deixar de ver um bloqueio que a tela
    mostra: os dois têm que listar as mesmas pendências.
    """
    import copy

    import openpyxl

    from apurabot.apuracao import apurar
    from apurabot.saida import escrever

    # Tira um CFOP do mapa de MS para que uma linha real fique sem atividade.
    p = copy.deepcopy(parametros)
    industriais = p.regimes["atividades"]["ms"]["por_cfop"]["industrial"]
    industriais["credito"] = [c for c in industriais["credito"] if c != 3101]

    apuracao = apurar(base_julho, p)
    assert apuracao.sem_regra_de_atividade, "o cenário não produziu a pendência"

    destino = tmp_path / "com_pendencia.xlsx"
    escrever(base_julho, destino, apuracao)
    texto = "\n".join(
        str(c.value)
        for linha in openpyxl.load_workbook(destino)["PENDÊNCIAS"].iter_rows()
        for c in linha
        if c.value
    )
    assert "ATIVIDADE INDEFINIDA" in texto
    assert "HINOVE (RIO BRILHANTE)" in texto
    assert "regimes.yaml" in texto


def test_o_resumo_diz_o_periodo_que_o_livro_cobre(base_julho):
    """Livro fechado ou pré-livro: quem lê o resultado precisa saber até onde vai."""
    assert base_julho.periodo == "01/07/2026 a 31/07/2026"
    assert base_julho.resumo()["periodo"] == base_julho.periodo


# -- o que a rodada de agosto pediu -----------------------------------------

def _avaliar(aba, coluna: str, linha: int) -> float:
    """Avalia as fórmulas que a aba escreve — SUM, subtração e ROUND.

    openpyxl grava fórmula, não resultado: para provar que a planilha entrega
    o mesmo número que o motor, é preciso resolvê-las. São três formas, todas
    escritas por `conferencia.py`, e nenhuma delas depende do Excel.
    """
    valor = aba[f"{coluna}{linha}"].value
    if isinstance(valor, (int, float)):
        return float(valor)
    if not isinstance(valor, str) or not valor.startswith("="):
        return 0.0

    achado = re.fullmatch(r"=SUM\((.+)\)", valor)
    if achado:
        total = 0.0
        for parte in achado.group(1).split(","):
            if ":" in parte:
                inicio, fim = parte.split(":")
                total += sum(
                    _avaliar(aba, inicio[0], n)
                    for n in range(int(inicio[1:]), int(fim[1:]) + 1)
                )
            else:
                total += _avaliar(aba, parte[0], int(parte[1:]))
        return total

    achado = re.fullmatch(r"=([A-Z])(\d+)-([A-Z])(\d+)", valor)
    if achado:
        a, b, c, d = achado.groups()
        return _avaliar(aba, a, int(b)) - _avaliar(aba, c, int(d))

    raise AssertionError(f"fórmula não reconhecida em {coluna}{linha}: {valor}")


@pytest.fixture(scope="module")
def com_formulas(base_julho, apuracao, tmp_path_factory):
    """A mesma planilha, lida com as fórmulas em vez do valor calculado."""
    destino = tmp_path_factory.mktemp("formulas") / "conferencia.xlsx"
    escrever(base_julho, destino, apuracao)
    return openpyxl.load_workbook(destino, data_only=False)


def test_os_totais_da_aba_sao_formula_e_batem_com_o_motor(com_formulas, apuracao):
    """Os totais deixaram de ser valor colado: agora somam as próprias linhas.

    Quem confere precisa ver de onde cada total veio, e a planilha precisa
    recalcular sozinha se alguém mexer numa linha. O que a fórmula resolve tem
    que ser exatamente o que o motor apurou — é o que este teste prova, filial
    por filial.
    """
    aba = com_formulas["APURAÇÃO EFETIVA"]
    estabelecimento, conferidas = None, 0
    for linha in aba.iter_rows(min_col=1, max_col=1):
        texto = str(linha[0].value or "")
        if "  —  " in texto:
            estabelecimento = texto.split("  —  ")[0]
        if texto != "TOTAL" or estabelecimento is None:
            continue
        n = linha[0].row
        filial = apuracao.filiais[estabelecimento]
        assert str(aba[f"E{n}"].value).startswith("=SUM("), estabelecimento
        assert _avaliar(aba, "G", n) == pytest.approx(
            filial.credito_bruto, abs=CENTAVO
        ), estabelecimento
        # A coluna soma o que não vira crédito: estorno da regra mais o
        # crédito indevido, que fica em parcela própria na apuração.
        assert _avaliar(aba, "J", n) == pytest.approx(
            filial.estorno + filial.credito_indevido, abs=CENTAVO
        ), estabelecimento
        assert _avaliar(aba, "K", n) == pytest.approx(
            filial.credito_mantido, abs=CENTAVO
        ), estabelecimento
        # A identidade que a coluna CHECK mostrava, conferida aqui em vez de
        # ocupar uma célula na planilha.
        assert _avaliar(aba, "J", n) + _avaliar(aba, "K", n) == pytest.approx(
            _avaliar(aba, "G", n), abs=CENTAVO
        ), estabelecimento
        conferidas += 1
    # Filial sem crédito de entrada não tem tabela — e por isso não tem TOTAL.
    com_credito = [f for f in apuracao.filiais.values() if f.credito_bruto]
    assert conferidas == len(com_credito)


def test_o_fechamento_traz_a_carga_efetiva_de_cada_classificacao(planilha):
    """Cada linha do bloco CRÉDITOS diz em que carga aquela classificação está."""
    aba = planilha["APURAÇÃO EFETIVA"]
    cabecalhos = [
        linha for linha in aba.iter_rows(max_col=7, values_only=True)
        if linha[1] == "Classificação" and linha[2] == "Carga efetiva"
    ]
    assert cabecalhos, "o bloco CRÉDITOS não tem a coluna de carga efetiva"

    # Matéria-prima em Guará entra toda a 4%: a média ponderada tem que ser 4%.
    # Só o bloco CRÉDITOS: o de DÉBITOS não tem coluna de estorno nem de carga.
    linhas = [
        linha for linha in aba.iter_rows(max_col=7, values_only=True)
        if linha[1] == "Matéria-Prima" and linha[4] and linha[5] is not None
    ]
    assert linhas
    assert all(l[2] == pytest.approx(0.04, abs=1e-6) for l in linhas), linhas


def test_os_blocos_nao_ficam_colados(planilha, apuracao):
    """Duas linhas em branco antes de cada estabelecimento."""
    aba = planilha["APURAÇÃO EFETIVA"]
    coluna = [linha[0].value for linha in aba.iter_rows(max_col=1)]
    titulos = [i for i, v in enumerate(coluna) if "  —  " in str(v or "")]
    assert len(titulos) == len(apuracao.filiais)
    for i in titulos:
        assert coluna[i - 1] is None and coluna[i - 2] is None, i
